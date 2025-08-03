import os
import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
from tkinterdnd2 import TkinterDnD, DND_FILES
from fpdf import FPDF
import webbrowser
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np

class FileSizeAnalyzer:
    def __init__(self):
        self.root = TkinterDnD.Tk()
        self.root.title("📁 File Size Analyzer")
        self.root.geometry("800x700")
        self.root.resizable(True, True)
        self.colors = {
            'bg': '#1e1e1e',
            'fg': '#ffffff',
            'secondary_bg': '#2d2d2d',
            'accent': '#007acc',
            'success': '#28a745',
            'warning': '#ffc107',
            'error': '#dc3545',
            'border': '#404040',
            'text_secondary': '#b0b0b0'
        }
        self.root.configure(bg=self.colors['bg'])
        self.selected_folder = None
        self.file_type_filter = []
        self.file_categories = {
            'Documents': ['.pdf', '.doc', '.docx', '.txt', '.rtf', '.odt'],
            'Images': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.svg'],
            'Videos': ['.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.webm'],
            'Audio': ['.mp3', '.wav', '.flac', '.aac', '.ogg', '.wma'],
            'Archives': ['.zip', '.rar', '.7z', '.tar', '.gz', '.bz2'],
            'Code': ['.py', '.js', '.html', '.css', '.java', '.cpp', '.c', '.php']
        }
        self.size_filter_enabled = False
        self.min_size_gb = 0.0
        self.max_size_gb = 1000.0
        self.date_filter_enabled = False
        self.start_date = None
        self.end_date = None
        self.search_enabled = False
        self.search_term = ""
        self.compression_algorithms = {
            'Images': {
                'JPEG Quality 85%': 0.3, 
                'WebP Conversion': 0.5,  
                'PNG Optimization': 0.2, 
                'AVIF Conversion': 0.6   
            },
            'Videos': {
                'H.264 Compression': 0.4,
                'H.265/HEVC': 0.6,       
                'AV1 Codec': 0.7,        
                'WebM Conversion': 0.5   
            },
            'Documents': {
                'PDF Compression': 0.3,  
                'ZIP Archive': 0.4,      
                '7-Zip Archive': 0.5     
            },
            'Audio': {
                'MP3 128kbps': 0.6,      
                'AAC 128kbps': 0.7,      
                'OGG Vorbis': 0.5        
            },
            'Archives': {
                '7-Zip LZMA2': 0.3,      
                'ZIP Deflate': 0.2,      
                'RAR5': 0.25             
            },
            'Code': {
                'Minification': 0.4,     
                'Gzip Compression': 0.3, 
                'Brotli Compression': 0.35 
            }
        }
        self.setup_ui()
        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind('<<Drop>>', self.handle_drop)
        
    def setup_ui(self):
        self.root.minsize(750, 650)
        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=0)
        self.root.grid_rowconfigure(2, weight=0)
        self.root.grid_rowconfigure(3, weight=0)
        self.root.grid_rowconfigure(4, weight=0)
        self.root.grid_rowconfigure(5, weight=0)
        self.root.grid_rowconfigure(6, weight=0)
        self.root.grid_rowconfigure(7, weight=0)
        self.root.grid_rowconfigure(8, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        title_label = tk.Label(
            self.root,
            text="📁 File Size Analyzer",
            font=("Segoe UI", 16, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['fg']
        )
        title_label.grid(row=0, column=0, pady=(10, 3), sticky="ew")
        desc_label = tk.Label(
            self.root,
            text="Analyze files and folders - Generate detailed size report",
            font=("Segoe UI", 10),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary']
        )
        desc_label.grid(row=1, column=0, pady=(3, 8), sticky="ew")
        self.folder_frame = tk.Frame(self.root, bg=self.colors['bg'])
        self.folder_frame.grid(row=2, column=0, pady=10, padx=20, sticky="ew")
        self.folder_frame.grid_columnconfigure(0, weight=1)
        self.folder_label = tk.Label(
            self.folder_frame,
            text="No folder selected",
            font=("Segoe UI", 10),
            bg=self.colors['secondary_bg'],
            fg=self.colors['text_secondary'],
            relief='flat',
            padx=8,
            pady=4,
            borderwidth=1,
            highlightthickness=1,
            highlightbackground=self.colors['border']
        )
        self.folder_label.grid(row=0, column=0, sticky="ew")
        self.filters_frame = tk.Frame(self.root, bg=self.colors['bg'])
        self.filters_frame.grid(row=3, column=0, pady=12, padx=25, sticky="ew")
        self.filters_frame.grid_columnconfigure(0, weight=1)
        filter_label = tk.Label(
            self.filters_frame,
            text="📁 File Type Filter:",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['fg']
        )
        filter_label.grid(row=0, column=0, sticky="w", pady=(0, 6))
        filter_buttons_frame = tk.Frame(self.filters_frame, bg=self.colors['bg'])
        filter_buttons_frame.grid(row=1, column=0, sticky="ew")
        self.filter_buttons = {}
        categories = list(self.file_categories.keys())
        
        for i, category in enumerate(categories):
            row_idx = i // 4
            col_idx = i % 4
            btn = tk.Button(
                filter_buttons_frame,
                text=f"📄 {category}",
                command=lambda cat=category: self.toggle_filter(cat),
                font=("Segoe UI", 10),
                bg=self.colors['secondary_bg'],
                fg=self.colors['text_secondary'],
                relief='flat',
                padx=10,
                pady=4,
                cursor='hand2',
                activebackground=self.colors['accent'],
                activeforeground=self.colors['fg']
            )
            btn.grid(row=row_idx, column=col_idx, padx=3, pady=3, sticky="ew")
            self.filter_buttons[category] = btn
        
        custom_frame = tk.Frame(self.filters_frame, bg=self.colors['bg'])
        custom_frame.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        custom_frame.grid_columnconfigure(0, weight=1)
        custom_label = tk.Label(
            custom_frame,
            text="Custom extensions:",
            font=("Segoe UI", 8),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary']
        )
        custom_label.grid(row=0, column=0, sticky="w", pady=(0, 2))
        self.custom_filter_entry = tk.Entry(
            custom_frame,
            font=("Segoe UI", 8),
            bg=self.colors['secondary_bg'],
            fg=self.colors['fg'],
            relief='flat',
            insertbackground=self.colors['fg']
        )
        self.custom_filter_entry.grid(row=1, column=0, sticky="ew", pady=(0, 4))
        self.custom_filter_entry.insert(0, "")
        self.custom_filter_entry.bind('<KeyRelease>', lambda e: self._update_status_message())
        clear_filter_btn = tk.Button(
            custom_frame,
            text="🗑️ Clear All Filters",
            command=self.clear_filters,
            font=("Segoe UI", 8),
            bg=self.colors['error'],
            fg=self.colors['fg'],
            relief='flat',
            padx=8,
            pady=2,
            cursor='hand2',
            activebackground='#c82333',
            activeforeground=self.colors['fg']
        )
        clear_filter_btn.grid(row=1, column=1, sticky="e", padx=(8, 0), pady=(0, 4))
        size_date_frame = tk.Frame(self.filters_frame, bg=self.colors['bg'])
        size_date_frame.grid(row=3, column=0, sticky="ew", pady=(8, 0))
        size_date_frame.grid_columnconfigure(0, weight=1)
        size_date_frame.grid_columnconfigure(1, weight=1)
        size_filter_label = tk.Label(
            size_date_frame,
            text="📏 Size Filter:",
            font=("Segoe UI", 9, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['fg']
        )
        size_filter_label.grid(row=0, column=0, sticky="w", pady=(0, 4))
        size_controls_frame = tk.Frame(size_date_frame, bg=self.colors['bg'])
        size_controls_frame.grid(row=1, column=0, sticky="ew")
        size_controls_frame.grid_columnconfigure(1, weight=1)
        size_controls_frame.grid_columnconfigure(3, weight=1)
        self.size_filter_var = tk.BooleanVar()
        self.size_filter_checkbox = tk.Checkbutton(
            size_controls_frame,
            text="Enable",
            variable=self.size_filter_var,
            command=self.toggle_size_filter,
            font=("Segoe UI", 8),
            bg=self.colors['bg'],
            fg=self.colors['fg'],
            selectcolor=self.colors['secondary_bg'],
            activebackground=self.colors['bg'],
            activeforeground=self.colors['fg']
        )
        self.size_filter_checkbox.grid(row=0, column=0, sticky="w", pady=(0, 6))
        min_label = tk.Label(
            size_controls_frame,
            text="Min (GB):",
            font=("Segoe UI", 8),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary']
        )
        min_label.grid(row=1, column=0, sticky="w", pady=(0, 2))
        self.min_size_entry = tk.Entry(
            size_controls_frame,
            font=("Segoe UI", 8),
            bg=self.colors['secondary_bg'],
            fg=self.colors['fg'],
            relief='flat',
            insertbackground=self.colors['fg'],
            width=10
        )
        self.min_size_entry.grid(row=1, column=1, sticky="ew", padx=(4, 0), pady=(0, 2))
        self.min_size_entry.insert(0, "0.0")
        self.min_size_entry.bind('<KeyRelease>', lambda e: self._update_status_message())
        max_label = tk.Label(
            size_controls_frame,
            text="Max (GB):",
            font=("Segoe UI", 8),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary']
        )
        max_label.grid(row=1, column=2, sticky="w", padx=(8, 0), pady=(0, 2))
        self.max_size_entry = tk.Entry(
            size_controls_frame,
            font=("Segoe UI", 8),
            bg=self.colors['secondary_bg'],
            fg=self.colors['fg'],
            relief='flat',
            insertbackground=self.colors['fg'],
            width=10
        )
        self.max_size_entry.grid(row=1, column=3, sticky="ew", padx=(4, 0), pady=(0, 2))
        self.max_size_entry.insert(0, "1000.0")
        self.max_size_entry.bind('<KeyRelease>', lambda e: self._update_status_message())
        presets_frame = tk.Frame(size_date_frame, bg=self.colors['bg'])
        presets_frame.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        presets = [
            ("Small (<1MB)", 0.0, 0.001),
            ("Medium (1MB-100MB)", 0.001, 0.1),
            ("Large (100MB-1GB)", 0.1, 1.0),
            ("Very Large (>1GB)", 1.0, 1000.0)
        ]
        
        for i, (text, min_val, max_val) in enumerate(presets):
            btn = tk.Button(
                presets_frame,
                text=text,
                command=lambda min_v=min_val, max_v=max_val: self.apply_size_preset(min_v, max_v),
                font=("Segoe UI", 7),
                bg=self.colors['secondary_bg'],
                fg=self.colors['text_secondary'],
                relief='flat',
                padx=6,
                pady=2,
                cursor='hand2',
                activebackground=self.colors['accent'],
                activeforeground=self.colors['fg']
            )
            btn.grid(row=0, column=i, padx=2, sticky="ew")

        date_filter_label = tk.Label(
            size_date_frame,
            text="📅 Date Filter:",
            font=("Segoe UI", 9, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['fg']
        )
        date_filter_label.grid(row=0, column=1, sticky="w", pady=(0, 4), padx=(20, 0))
        date_controls_frame = tk.Frame(size_date_frame, bg=self.colors['bg'])
        date_controls_frame.grid(row=1, column=1, sticky="ew", padx=(20, 0))
        date_controls_frame.grid_columnconfigure(1, weight=1)
        date_controls_frame.grid_columnconfigure(3, weight=1)
        self.date_filter_var = tk.BooleanVar()
        self.date_filter_checkbox = tk.Checkbutton(
            date_controls_frame,
            text="Enable",
            variable=self.date_filter_var,
            command=self.toggle_date_filter,
            font=("Segoe UI", 8),
            bg=self.colors['bg'],
            fg=self.colors['fg'],
            selectcolor=self.colors['secondary_bg'],
            activebackground=self.colors['bg'],
            activeforeground=self.colors['fg']
        )
        self.date_filter_checkbox.grid(row=0, column=0, sticky="w", pady=(0, 6))
        start_date_label = tk.Label(
            date_controls_frame,
            text="Start:",
            font=("Segoe UI", 8),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary']
        )
        start_date_label.grid(row=1, column=0, sticky="w", pady=(0, 2))
        self.start_date_entry = tk.Entry(
            date_controls_frame,
            font=("Segoe UI", 8),
            bg=self.colors['secondary_bg'],
            fg=self.colors['fg'],
            relief='flat',
            insertbackground=self.colors['fg'],
            width=12
        )
        self.start_date_entry.grid(row=1, column=1, sticky="ew", padx=(4, 0), pady=(0, 2))
        self.start_date_entry.insert(0, "2020-01-01")
        self.start_date_entry.bind('<KeyRelease>', lambda e: self._update_status_message())
        end_date_label = tk.Label(
            date_controls_frame,
            text="End:",
            font=("Segoe UI", 8),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary']
        )
        end_date_label.grid(row=1, column=2, sticky="w", padx=(8, 0), pady=(0, 2))
        self.end_date_entry = tk.Entry(
            date_controls_frame,
            font=("Segoe UI", 8),
            bg=self.colors['secondary_bg'],
            fg=self.colors['fg'],
            relief='flat',
            insertbackground=self.colors['fg'],
            width=12
        )
        self.end_date_entry.grid(row=1, column=3, sticky="ew", padx=(4, 0), pady=(0, 2))
        self.end_date_entry.insert(0, "2025-12-31")
        self.end_date_entry.bind('<KeyRelease>', lambda e: self._update_status_message())
        date_presets_frame = tk.Frame(size_date_frame, bg=self.colors['bg'])
        date_presets_frame.grid(row=2, column=1, sticky="ew", pady=(6, 0), padx=(20, 0))
        date_presets = [
            ("7 Days", "7"),
            ("30 Days", "30"),
            ("3 Months", "90"),
            ("1 Year", "365")
        ]
        
        for i, (text, days) in enumerate(date_presets):
            btn = tk.Button(
                date_presets_frame,
                text=text,
                command=lambda d=days: self.apply_date_preset(d),
                font=("Segoe UI", 7),
                bg=self.colors['secondary_bg'],
                fg=self.colors['text_secondary'],
                relief='flat',
                padx=6,
                pady=2,
                cursor='hand2',
                activebackground=self.colors['accent'],
                activeforeground=self.colors['fg']
            )
            btn.grid(row=0, column=i, padx=2, sticky="ew")

        search_frame = tk.Frame(self.filters_frame, bg=self.colors['bg'])
        search_frame.grid(row=4, column=0, sticky="ew", pady=(8, 0))
        search_frame.grid_columnconfigure(1, weight=1)
        search_label = tk.Label(
            search_frame,
            text="🔍 Search Filter:",
            font=("Segoe UI", 9, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['fg']
        )
        search_label.grid(row=0, column=0, sticky="w", pady=(0, 4))
        search_controls_frame = tk.Frame(search_frame, bg=self.colors['bg'])
        search_controls_frame.grid(row=1, column=0, columnspan=2, sticky="ew")
        search_controls_frame.grid_columnconfigure(1, weight=1)
        self.search_filter_var = tk.BooleanVar()
        self.search_filter_checkbox = tk.Checkbutton(
            search_controls_frame,
            text="Enable",
            variable=self.search_filter_var,
            command=self.toggle_search_filter,
            font=("Segoe UI", 8),
            bg=self.colors['bg'],
            fg=self.colors['fg'],
            selectcolor=self.colors['secondary_bg'],
            activebackground=self.colors['bg'],
            activeforeground=self.colors['fg']
        )
        self.search_filter_checkbox.grid(row=0, column=0, sticky="w", pady=(0, 6))
        search_entry_label = tk.Label(
            search_controls_frame,
            text="Search term:",
            font=("Segoe UI", 8),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary']
        )
        search_entry_label.grid(row=1, column=0, sticky="w", pady=(0, 2))
        self.search_entry = tk.Entry(
            search_controls_frame,
            font=("Segoe UI", 8),
            bg=self.colors['secondary_bg'],
            fg=self.colors['fg'],
            relief='flat',
            insertbackground=self.colors['fg']
        )
        self.search_entry.grid(row=1, column=1, sticky="ew", padx=(4, 0), pady=(0, 2))
        self.search_entry.insert(0, "Enter file or folder name...")
        self.search_entry.bind('<KeyRelease>', lambda e: self._update_status_message())
        self.search_entry.bind('<FocusIn>', self.on_search_focus_in)
        self.search_entry.bind('<FocusOut>', self.on_search_focus_out)
        search_options_frame = tk.Frame(search_frame, bg=self.colors['bg'])
        search_options_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        self.search_case_sensitive = tk.BooleanVar()
        self.search_case_checkbox = tk.Checkbutton(
            search_options_frame,
            text="Case sensitive",
            variable=self.search_case_sensitive,
            font=("Segoe UI", 7),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary'],
            selectcolor=self.colors['secondary_bg'],
            activebackground=self.colors['bg'],
            activeforeground=self.colors['text_secondary']
        )
        self.search_case_checkbox.grid(row=0, column=0, sticky="w")
        self.search_exact_match = tk.BooleanVar()
        self.search_exact_checkbox = tk.Checkbutton(
            search_options_frame,
            text="Exact match",
            variable=self.search_exact_match,
            font=("Segoe UI", 7),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary'],
            selectcolor=self.colors['secondary_bg'],
            activebackground=self.colors['bg'],
            activeforeground=self.colors['text_secondary']
        )
        self.search_exact_checkbox.grid(row=0, column=1, sticky="w", padx=(20, 0))
        button_frame = tk.Frame(self.root, bg=self.colors['bg'])
        button_frame.grid(row=5, column=0, pady=10, sticky="ew")
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)
        button_frame.grid_columnconfigure(2, weight=1)
        button_frame.grid_columnconfigure(3, weight=1)
        button_frame.grid_columnconfigure(4, weight=1)
        button_frame.grid_columnconfigure(5, weight=1)
        self.select_button = tk.Button(
            button_frame,
            text="📂 Select Folder",
            command=self.select_folder,
            font=("Segoe UI", 10, "bold"),
            bg=self.colors['accent'],
            fg=self.colors['fg'],
            relief='flat',
            padx=12,
            pady=6,
            cursor='hand2',
            activebackground='#005a9e',
            activeforeground=self.colors['fg']
        )
        self.select_button.grid(row=0, column=0, padx=6, sticky="ew")
        self.export_button = tk.Button(
            button_frame,
            text="📊 Excel",
            command=self.export_to_excel,
            font=("Segoe UI", 10, "bold"),  
            bg=self.colors['success'],
            fg=self.colors['fg'],
            relief='flat',
            padx=12,
            pady=6,
            cursor='hand2',
            state='disabled',
            activebackground='#1e7e34',
            activeforeground=self.colors['fg']
        )
        self.export_button.grid(row=0, column=1, padx=6, sticky="ew")
        self.pdf_button = tk.Button(
            button_frame,
            text="📄 PDF",
            command=self.export_to_pdf,
            font=("Segoe UI", 10, "bold"),
            bg=self.colors['warning'],
            fg=self.colors['fg'],
            relief='flat',
            padx=12,
            pady=6,
            cursor='hand2',
            state='disabled',
            activebackground='#e0a800',
            activeforeground=self.colors['fg']
        )
        self.pdf_button.grid(row=0, column=2, padx=6, sticky="ew")
        self.html_button = tk.Button(
            button_frame,
            text="🌐 HTML",
            command=self.export_to_html,
            font=("Segoe UI", 10, "bold"),
            bg=self.colors['error'],
            fg=self.colors['fg'],
            relief='flat',
            padx=12,
            pady=6,
            cursor='hand2',
            state='disabled',
            activebackground='#c82333',
            activeforeground=self.colors['fg']
        )
        self.html_button.grid(row=0, column=3, padx=6, sticky="ew")
        self.charts_button = tk.Button(
            button_frame,
            text="📈 Charts",
            command=self.show_charts,
            font=("Segoe UI", 10, "bold"),
            bg='#6f42c1',
            fg=self.colors['fg'],
            relief='flat',
            padx=12,
            pady=6,
            cursor='hand2',
            state='disabled',
            activebackground='#5a32a3',
            activeforeground=self.colors['fg']
        )
        self.charts_button.grid(row=0, column=4, padx=6, sticky="ew")
        self.optimize_button = tk.Button(
            button_frame,
            text="💾 Optimize",
            command=self.show_optimization,
            font=("Segoe UI", 10, "bold"),
            bg='#fd7e14',
            fg=self.colors['fg'],
            relief='flat',
            padx=12,
            pady=6,
            cursor='hand2',
            state='disabled',
            activebackground='#e8690b',
            activeforeground=self.colors['fg']
        )
        self.optimize_button.grid(row=0, column=5, padx=6, sticky="ew")
        self.progress = ttk.Progressbar(
            self.root,
            mode='indeterminate',
            length=300
        )
        self.progress.grid(row=6, column=0, pady=10, sticky="ew")
        self.status_label = tk.Label(
            self.root,
            text="Select a folder and start analysis",
            font=("Segoe UI", 11),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary']
        )
        self.status_label.grid(row=7, column=0, pady=8, sticky="ew")
        self.result_frame = tk.Frame(self.root, bg=self.colors['bg'])
        self.result_frame.grid(row=8, column=0, pady=10, padx=25, sticky="nsew")
        self.result_frame.grid_columnconfigure(0, weight=1)
        self.result_label = tk.Label(
            self.result_frame,
            text="",
            font=("Segoe UI", 11),
            bg=self.colors['bg'],
            fg=self.colors['fg'],
            justify='left',
            anchor='nw'
        )
        self.result_label.grid(row=0, column=0, sticky="nsew")
        self.selected_folder = None
        self.files_data = []
        self.file_type_filter = []
        self.file_categories = {
            'Documents': ['.pdf', '.doc', '.docx', '.txt', '.rtf', '.odt', '.pages'],
            'Images': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.svg', '.webp'],
            'Videos': ['.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.webm', '.m4v'],
            'Audio': ['.mp3', '.wav', '.flac', '.aac', '.ogg', '.wma', '.m4a'],
            'Archives': ['.zip', '.rar', '.7z', '.tar', '.gz', '.bz2', '.xz'],
            'Code': ['.py', '.js', '.html', '.css', '.java', '.cpp', '.c', '.php', '.sql'],
            'Spreadsheets': ['.xlsx', '.xls', '.csv', '.ods', '.numbers'],
            'Presentations': ['.pptx', '.ppt', '.odp', '.key']
        }
        
        self.size_filter_enabled = False
        self.min_size_gb = 0.0
        self.max_size_gb = 1000.0
        
    def select_folder(self):
        # Sadece klasör seçimi yapılacak
        folder_path = filedialog.askdirectory(
            title="Select folder to analyze",
            initialdir=os.path.expanduser("~")
        )
        
        if folder_path:
            self.selected_folder = folder_path
            display_text = f"Selected Folder: {folder_path}"
            
            self.folder_label.config(
                text=display_text,
                fg=self.colors['success']
            )
            self.export_button.config(state='normal')
            self.pdf_button.config(state='normal')
            self.html_button.config(state='normal')
            self.charts_button.config(state='normal')
            self.optimize_button.config(state='normal')
            self._update_status_message()
            
    def handle_drop(self, event):
        dropped_path = event.data.strip('{}')
        if os.path.isdir(dropped_path):
            self.selected_folder = dropped_path
            display_text = f"Selected Folder: {dropped_path}"
            self.folder_label.config(
                text=display_text,
                fg=self.colors['success']
            )
            self.export_button.config(state='normal')
            self.pdf_button.config(state='normal')
            self.html_button.config(state='normal')
            self.charts_button.config(state='normal')
            self.optimize_button.config(state='normal')
            self._update_status_message()
            self.status_label.config(text="Folder selected via drag & drop. " + self.status_label.cget("text").replace("Folder selected. ", ""))
        else:
            from tkinter import messagebox
            messagebox.showwarning("Warning", "Please drop a folder, not a file!")
            
    def toggle_filter(self, category):

        if category in self.file_type_filter:
            self.file_type_filter.remove(category)
            self.filter_buttons[category].config(
                bg=self.colors['secondary_bg'],
                fg=self.colors['text_secondary']
            )
        else:
            self.file_type_filter.append(category)
            self.filter_buttons[category].config(
                bg=self.colors['accent'],
                fg=self.colors['fg']
            )

        self._update_status_message()
    
    def clear_filters(self):
        self.file_type_filter.clear()
        for btn in self.filter_buttons.values():
            btn.config(
                bg=self.colors['secondary_bg'],
                fg=self.colors['text_secondary']
            )
        self.custom_filter_entry.delete(0, tk.END)
        self.custom_filter_entry.insert(0, "")
        self.size_filter_var.set(False)
        self.size_filter_enabled = False
        self.min_size_entry.delete(0, tk.END)
        self.min_size_entry.insert(0, "0.0")
        self.max_size_entry.delete(0, tk.END)
        self.max_size_entry.insert(0, "1000.0")
        self.date_filter_var.set(False)
        self.date_filter_enabled = False
        self.start_date_entry.delete(0, tk.END)
        self.start_date_entry.insert(0, "2020-01-01")
        self.end_date_entry.delete(0, tk.END)
        self.end_date_entry.insert(0, "2025-12-31")
        self.search_filter_var.set(False)
        self.search_enabled = False
        self.search_entry.delete(0, tk.END)
        self.search_entry.insert(0, "Enter file or folder name...")
        self.search_entry.config(fg=self.colors['text_secondary'])
        self.search_case_sensitive.set(False)
        self.search_exact_match.set(False)
        self._update_status_message()
    
    def get_active_filters(self):
        active_extensions = []
        
        for category in self.file_type_filter:
            active_extensions.extend(self.file_categories[category])

        custom_text = self.custom_filter_entry.get().strip()
        if custom_text:
            custom_extensions = [ext.strip() for ext in custom_text.split(',') if ext.strip()]
            active_extensions.extend(custom_extensions)
        
        unique_extensions = []
        for ext in active_extensions:
            if not ext.startswith('.'):
                ext = '.' + ext
            if ext not in unique_extensions:
                unique_extensions.append(ext)
        
        # Eğer hiç filtre seçilmemişse boş liste döndür (tüm dosyaları göster)
        return unique_extensions
    
    def toggle_size_filter(self):
        self.size_filter_enabled = self.size_filter_var.get()
        self._update_status_message()
    
    def apply_size_preset(self, min_size, max_size):
        self.size_filter_var.set(True)
        self.size_filter_enabled = True
        self.min_size_gb = min_size
        self.max_size_gb = max_size
        self.min_size_entry.delete(0, tk.END)
        self.min_size_entry.insert(0, str(min_size))
        self.max_size_entry.delete(0, tk.END)
        self.max_size_entry.insert(0, str(max_size))
        self._update_status_message()
    
    def get_size_filter_range(self):
        try:
            min_size = float(self.min_size_entry.get())
            max_size = float(self.max_size_entry.get())
            return min_size, max_size
        except ValueError:
            return 0.0, 1000.0
    
    def toggle_date_filter(self):
        self.date_filter_enabled = self.date_filter_var.get()
        self._update_status_message()
    
    def apply_date_preset(self, days):
        from datetime import datetime, timedelta
        
        self.date_filter_var.set(True)
        self.date_filter_enabled = True
        end_date = datetime.now()
        start_date = end_date - timedelta(days=int(days))
        self.start_date_entry.delete(0, tk.END)
        self.start_date_entry.insert(0, start_date.strftime("%Y-%m-%d"))
        self.end_date_entry.delete(0, tk.END)
        self.end_date_entry.insert(0, end_date.strftime("%Y-%m-%d"))
        self._update_status_message()
    
    def get_date_filter_range(self):
        try:
            from datetime import datetime
            start_date_str = self.start_date_entry.get()
            end_date_str = self.end_date_entry.get()
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            return start_date, end_date
        except ValueError:
            from datetime import datetime
            return datetime(2020, 1, 1), datetime(2025, 12, 31)
    
    def toggle_search_filter(self):
        self.search_enabled = self.search_filter_var.get()
        self._update_status_message()
    
    def on_search_focus_in(self, event):
        if self.search_entry.get() == "Enter file or folder name...":
            self.search_entry.delete(0, tk.END)
            self.search_entry.config(fg=self.colors['fg'])
    
    def on_search_focus_out(self, event):
        if not self.search_entry.get().strip():
            self.search_entry.insert(0, "Enter file or folder name...")
            self.search_entry.config(fg=self.colors['text_secondary'])
    
    def get_search_term(self):
        search_term = self.search_entry.get().strip()
        if search_term == "Enter file or folder name...":
            return ""
        return search_term
    
    def matches_search(self, item_name):
        if not self.search_enabled:
            return True
        
        search_term = self.get_search_term()
        if not search_term:
            return True
        
        case_sensitive = self.search_case_sensitive.get()
        exact_match = self.search_exact_match.get()
        
        if case_sensitive:
            item_str = item_name
            search_str = search_term
        else:
            item_str = item_name.lower()
            search_str = search_term.lower()
        
        if exact_match:
            return item_str == search_str
        else:
            return search_str in item_str
    
    def _update_status_message(self):
        if not self.selected_folder:
            return
            
        active_filters = self.get_active_filters()
        size_info = ""
        date_info = ""
        search_info = ""
        
        if self.size_filter_enabled:
            min_size, max_size = self.get_size_filter_range()
            size_info = f"Size: {min_size}-{max_size}GB"
        
        if self.date_filter_enabled:
            start_date, end_date = self.get_date_filter_range()
            date_info = f"Date: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        
        if self.search_enabled:
            search_term = self.get_search_term()
            if search_term:
                search_info = f"Search: '{search_term}'"
        
        if active_filters or self.size_filter_enabled or self.date_filter_enabled or self.search_enabled:
            filter_parts = []
            if active_filters:
                filter_parts.append(f"Types: {', '.join(active_filters[:3])}{'...' if len(active_filters) > 3 else ''}")
            if size_info:
                filter_parts.append(size_info)
            if date_info:
                filter_parts.append(date_info)
            if search_info:
                filter_parts.append(search_info)
            
            filter_text = " | ".join(filter_parts)
            self.status_label.config(text=f"Folder selected. {filter_text}. Click button to create reports.")
        else:
            self.status_label.config(text="Folder selected. No filters active. Click button to create reports.")
            
    def get_folder_size(self, folder_path):
        total_size = 0
        try:
            for item in Path(folder_path).rglob('*'):
                if item.is_file():
                    total_size += item.stat().st_size
        except PermissionError:
            pass
        return total_size
    
    def analyze_folder_contents(self, folder_path, parent_folder=""):
        files_data = []
        
        try:
            selected_path = Path(folder_path)
            system_folders = ['$RECYCLE.BIN', 'System Volume Information', 'RECYCLER', 'Thumbs.db']
            active_filters = self.get_active_filters()
            size_filter_enabled = self.size_filter_enabled
            min_size_gb, max_size_gb = self.get_size_filter_range()
            date_filter_enabled = self.date_filter_enabled
            start_date, end_date = self.get_date_filter_range()
            search_filter_enabled = self.search_enabled
            
            # Sadece klasör analizi yapılacak
            if not selected_path.is_dir():
                return [{
                    'Name': 'Error: Not a folder',
                    'Type': 'Error',
                    'Size (GB)': 0,
                    'Extension': '❌',
                    'Full Path': str(selected_path)
                }]
            
            # Seçilen klasörün doğrudan altındaki öğeleri analiz et
            for item in selected_path.iterdir():
                # Sistem klasörlerini atla
                if item.name in system_folders:
                    continue
                
                if item.is_dir():
                    # Alt klasörün boyutunu hesapla
                    folder_size = self.get_folder_size(item)
                    folder_size_mb = folder_size / (1024 * 1024)
                    folder_size_gb = folder_size_mb / 1024
                    
                    # Arama filtresini kontrol et
                    folder_passed_search_filter = True
                    if search_filter_enabled:
                        folder_passed_search_filter = self.matches_search(item.name)
                    
                    if folder_passed_search_filter:
                        files_data.append({
                            'Name': item.name,
                            'Type': 'Folder',
                            'Size (GB)': round(folder_size_gb, 2),
                            'Extension': '📁',
                            'Full Path': str(item)
                        })
                
                elif item.is_file():
                    # Dosya filtrelerini kontrol et
                    file_passed_type_filter = True
                    if active_filters:
                        file_extension = item.suffix.lower()
                        file_passed_type_filter = file_extension in active_filters
                    
                    file_passed_size_filter = True
                    if size_filter_enabled:
                        size_bytes = item.stat().st_size
                        size_gb = size_bytes / (1024 * 1024 * 1024)
                        file_passed_size_filter = min_size_gb <= size_gb <= max_size_gb
                    
                    file_passed_date_filter = True
                    if date_filter_enabled:
                        from datetime import datetime
                        file_creation_time = datetime.fromtimestamp(item.stat().st_ctime)
                        file_passed_date_filter = start_date <= file_creation_time <= end_date
                    
                    file_passed_search_filter = True
                    if search_filter_enabled:
                        file_passed_search_filter = self.matches_search(item.name)
                    
                    # Tüm filtreleri geçerse dosyayı ekle
                    if file_passed_type_filter and file_passed_size_filter and file_passed_date_filter and file_passed_search_filter:
                        size_bytes = item.stat().st_size
                        size_mb = size_bytes / (1024 * 1024)
                        size_gb = size_mb / 1024
                        
                        files_data.append({
                            'Name': item.name,
                            'Type': 'File',
                            'Size (GB)': round(size_gb, 2),
                            'Extension': item.suffix.lower(),
                            'Full Path': str(item)
                        })
                    
                    # Debug için: Filtre geçmeyen dosyaları da göster
                    else:
                        print(f"File filtered out: {item.name} - Type: {file_passed_type_filter}, Size: {file_passed_size_filter}, Date: {file_passed_date_filter}, Search: {file_passed_search_filter}")
                        print(f"Active filters: {active_filters}")
                        print(f"File extension: {item.suffix.lower()}")
            
            return files_data
            
        except Exception as e:
            return [{
                'Name': f'Error: {str(e)}',
                'Type': 'Error',
                'Size (GB)': 0,
                'Extension': '❌',
                'Full Path': folder_path
            }]
    
    def get_file_sizes(self):
        try:
            self.progress.start()
            self.status_label.config(text="Analyzing folder contents and subfolders...")
            self.root.update()
            all_data = self.analyze_folder_contents(self.selected_folder)
            total_size_gb = sum(item['Size (GB)'] for item in all_data if item['Type'] in ['File', 'Folder'])
            self.progress.stop()
            return all_data, total_size_gb
            
        except Exception as e:
            self.progress.stop()
            raise e
    
    def export_to_excel(self):
        if not self.selected_folder:
            messagebox.showwarning("Warning", "Please select a folder first!")
            return
            
        try:
            self.files_data, total_size = self.get_file_sizes()
            
            if not self.files_data:
                messagebox.showinfo("Info", "No files or folders found in selected location!")
                return
            
            df = pd.DataFrame(self.files_data)
            total_size_gb = total_size
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            selected_folder_name = os.path.basename(self.selected_folder)
            drive_letter = os.path.splitdrive(self.selected_folder)[0]
            
            try:
                import shutil
                total, used, free = shutil.disk_usage(self.selected_folder)
                drive_size_gb = total // (1024**3)
                drive_name = f"{drive_letter} ({drive_size_gb}GB)"
            except:
                drive_name = drive_letter
            
            output_filename = f"{drive_name}_{selected_folder_name}_analysis_{timestamp}.xlsx"
            current_dir = os.path.dirname(os.path.abspath(__file__))
            output_path = os.path.join(current_dir, output_filename)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                workbook = writer.book
                folders_df = df[df['Type'] == 'Folder'].copy()
                folders_df = folders_df.sort_values(['Name'])
                files_df = df[df['Type'] == 'File'].copy()
                files_df = files_df.sort_values(['Name'])
                errors_df = df[df['Type'] == 'Error'].copy()
                all_data = pd.concat([folders_df, files_df, errors_df], ignore_index=True)
                worksheet = workbook.create_sheet('Detailed Analysis')
                title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
                header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
                normal_font = Font(name='Segoe UI', size=10)
                title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                folder_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                file_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                error_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                center_alignment = Alignment(horizontal='center', vertical='center')
                left_alignment = Alignment(horizontal='left', vertical='center')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                worksheet.merge_cells('A1:E1')
                title_cell = worksheet['A1']
                title_cell.value = f"📁 {drive_name} - {selected_folder_name} - Detailed File Analysis"
                title_cell.font = title_font
                title_cell.fill = title_fill
                title_cell.alignment = center_alignment
                title_cell.border = thin_border
                worksheet.merge_cells('A2:E2')
                subtitle_cell = worksheet['A2']
                subtitle_cell.value = f"📊 Analysis Date: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')} | 📍 Folder: {self.selected_folder}"
                subtitle_cell.font = Font(name='Segoe UI', size=9, color='666666')
                subtitle_cell.alignment = center_alignment
                headers = ['Name', 'Type', 'Size (GB)', 'Extension', 'Full Path']

                for col, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=4, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                row_num = 5
                for _, data_row in all_data.iterrows():
                    for col, value in enumerate(data_row, 1):
                        cell = worksheet.cell(row=row_num, column=col)
                        cell.value = value
                        cell.font = normal_font
                        cell.alignment = left_alignment
                        cell.border = thin_border
                        
                        if data_row['Type'] == 'Folder':
                            cell.fill = folder_fill
                        elif data_row['Type'] == 'File':
                            cell.fill = file_fill
                        elif data_row['Type'] == 'Error':
                            cell.fill = error_fill
                    
                    row_num += 1
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        if cell.coordinate in worksheet.merged_cells:
                            continue
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                                column_letter = cell.column_letter
                        except:
                            pass
                    if column_letter:
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                folder_summary = df.groupby('Name').agg({
                    'Type': 'count',
                    'Size (GB)': 'sum'
                }).reset_index()
                folder_summary.columns = ['Name', 'Item Count', 'Total Size (GB)']
                folder_summary['Total Size (GB)'] = folder_summary['Total Size (GB)'].round(2)
                folder_summary = folder_summary.sort_values('Total Size (GB)', ascending=False)
                worksheet2 = workbook.create_sheet('Folder Summary')
                worksheet2.merge_cells('A1:C1')
                title_cell2 = worksheet2['A1']
                title_cell2.value = f"📂 {drive_name} - {selected_folder_name} - Folder Summary"
                title_cell2.font = title_font
                title_cell2.fill = title_fill
                title_cell2.alignment = center_alignment
                title_cell2.border = thin_border
                headers2 = ['Name', 'Item Count', 'Total Size (GB)']

                for col, header in enumerate(headers2, 1):
                    cell = worksheet2.cell(row=3, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                for row_idx, (_, data_row) in enumerate(folder_summary.iterrows(), 4):
                    for col_idx, value in enumerate(data_row, 1):
                        cell = worksheet2.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.font = normal_font
                        cell.alignment = left_alignment
                        cell.border = thin_border
                        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                
                for column in worksheet2.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        if cell.coordinate in worksheet2.merged_cells:
                            continue
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                                column_letter = cell.column_letter
                        except:
                            pass
                    if column_letter:
                        adjusted_width = min(max_length + 2, 30)
                        worksheet2.column_dimensions[column_letter].width = adjusted_width
                
                files_count = len([item for item in self.files_data if item['Type'] == 'File'])
                folders_count = len([item for item in self.files_data if item['Type'] == 'Folder'])
                error_count = len([item for item in self.files_data if item['Type'] == 'Error'])
                summary_data = {
                    'Metric': [
                        'Total Items',
                        'File Count',
                        'Folder Count',
                        'Error Count',
                        'Total Size (GB)',
                        'Largest File',
                        'Smallest File',
                        'Average File Size (GB)',
                        'Analyzed Folders'
                    ],
                    'Value': [
                        len(self.files_data),
                        files_count,
                        folders_count,
                        error_count,
                        round(total_size_gb, 2),
                        max([item for item in self.files_data if item['Type'] == 'File'], 
                            key=lambda x: x['Size (GB)'])['Name'] if files_count > 0 else 'None',
                        min([item for item in self.files_data if item['Type'] == 'File'], 
                            key=lambda x: x['Size (GB)'])['Name'] if files_count > 0 else 'None',
                        round(total_size_gb / files_count, 2) if files_count > 0 else 0,
                        len(df['Name'].unique())
                    ]
                }
                
                summary_df = pd.DataFrame(summary_data)
                worksheet3 = workbook.create_sheet('General Summary')
                worksheet3.merge_cells('A1:B1')
                title_cell3 = worksheet3['A1']
                title_cell3.value = f"📊 {drive_name} - {selected_folder_name} - General Summary"
                title_cell3.font = title_font
                title_cell3.fill = title_fill
                title_cell3.alignment = center_alignment
                title_cell3.border = thin_border
                headers3 = ['Metric', 'Value']

                for col, header in enumerate(headers3, 1):
                    cell = worksheet3.cell(row=3, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                for row_idx, (_, data_row) in enumerate(summary_df.iterrows(), 4):
                    for col_idx, value in enumerate(data_row, 1):
                        cell = worksheet3.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.font = normal_font
                        cell.alignment = left_alignment
                        cell.border = thin_border
                        cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                
                for column in worksheet3.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        if cell.coordinate in worksheet3.merged_cells:
                            continue
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                                column_letter = cell.column_letter
                        except:
                            pass
                    if column_letter:
                        adjusted_width = min(max_length + 2, 40)
                        worksheet3.column_dimensions[column_letter].width = adjusted_width
                
                file_types = df[df['Type'] == 'File']['Extension'].value_counts().reset_index()
                file_types.columns = ['File Extension', 'Count']
                worksheet4 = workbook.create_sheet('File Types')
                worksheet4.merge_cells('A1:B1')
                title_cell4 = worksheet4['A1']
                title_cell4.value = f"📄 {drive_name} - {selected_folder_name} - File Types"
                title_cell4.font = title_font
                title_cell4.fill = title_fill
                title_cell4.alignment = center_alignment
                title_cell4.border = thin_border
                headers4 = ['File Extension', 'Count']

                for col, header in enumerate(headers4, 1):
                    cell = worksheet4.cell(row=3, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                for row_idx, (_, data_row) in enumerate(file_types.iterrows(), 4):
                    for col_idx, value in enumerate(data_row, 1):
                        cell = worksheet4.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.font = normal_font
                        cell.alignment = left_alignment
                        cell.border = thin_border
                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                
                for column in worksheet4.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        if cell.coordinate in worksheet4.merged_cells:
                            continue
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                                column_letter = cell.column_letter
                        except:
                            pass
                    if column_letter:
                        adjusted_width = min(max_length + 2, 25)
                        worksheet4.column_dimensions[column_letter].width = adjusted_width
                
                if 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])
            
            result_text = f"""
📊 Detailed Analysis Completed!

📁 Total Items: {len(self.files_data)}
📄 File Count: {files_count}
📂 Folder Count: {folders_count}
❌ Error Count: {error_count}
💾 Total Size: {round(total_size_gb, 2)} GB
📊 Analyzed Folders: {len(df['Name'].unique())}
📄 Excel File: {output_filename}
📍 Location: {current_dir}
            """
            
            self.result_label.config(text=result_text)
            self.status_label.config(text="Detailed Excel file created successfully!")
            
            messagebox.showinfo(
                "Success", 
                f"Detailed Excel file created!\n\n"
                f"File: {output_filename}\n"
                f"Location: {current_dir}\n\n"
                f"Total {len(self.files_data)} items analyzed.\n"
                f"({files_count} files, {folders_count} folders, {error_count} errors)\n"
                f"{len(df['Name'].unique())} folders analyzed in detail."
            )
            
        except Exception as e:
            self.progress.stop()
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
            self.status_label.config(text="Error occurred!")
    
    def export_to_pdf(self):
        if not self.selected_folder:
            messagebox.showwarning("Warning", "Please select a folder first!")
            return
            
        if not self.files_data:
            messagebox.showinfo("Info", "Please run analysis first by creating Excel report!")
            return
            
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            selected_folder_name = os.path.basename(self.selected_folder)
            drive_letter = os.path.splitdrive(self.selected_folder)[0]
            
            try:
                import shutil
                total, used, free = shutil.disk_usage(self.selected_folder)
                drive_size_gb = total // (1024**3)
                drive_name = f"{drive_letter} ({drive_size_gb}GB)"
            except:
                drive_name = drive_letter
            
            output_filename = f"{drive_name}_{selected_folder_name}_analysis_{timestamp}.pdf"
            current_dir = os.path.dirname(os.path.abspath(__file__))
            output_path = os.path.join(current_dir, output_filename)
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", 'B', 16)
            title = f"📁 {drive_name} - {selected_folder_name} - Detailed File Analysis"
            pdf.cell(0, 10, title, ln=True, align='C')
            pdf.set_font("Helvetica", '', 10)
            subtitle = f"📊 Analysis Date: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
            pdf.cell(0, 8, subtitle, ln=True, align='C')
            folder_info = f"📍 Folder: {self.selected_folder}"
            pdf.cell(0, 8, folder_info, ln=True, align='C')
            pdf.ln(10)
            files_count = len([item for item in self.files_data if item['Type'] == 'File'])
            folders_count = len([item for item in self.files_data if item['Type'] == 'Folder'])
            error_count = len([item for item in self.files_data if item['Type'] == 'Error'])
            total_size_gb = sum(item['Size (GB)'] for item in self.files_data if item['Type'] in ['File', 'Folder'])
            summary_text = f"📊 Summary: {len(self.files_data)} total items ({files_count} files, {folders_count} folders, {error_count} errors)"
            pdf.cell(0, 8, summary_text, ln=True)
            size_text = f"💾 Total Size: {round(total_size_gb, 2)} GB"
            pdf.cell(0, 8, size_text, ln=True)
            pdf.ln(10)
            pdf.set_font("Helvetica", 'B', 12)
            pdf.cell(0, 8, "Detailed File Analysis", ln=True)
            pdf.set_font("Helvetica", '', 9)
            headers = ['Name', 'Type', 'Size (GB)', 'Extension']
            col_widths = [60, 20, 25, 25]
            
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, align='C')
            pdf.ln()
            
            for item in self.files_data:
                pdf.cell(col_widths[0], 6, str(item['Name'])[:50], border=1)
                pdf.cell(col_widths[1], 6, str(item['Type']), border=1, align='C')
                pdf.cell(col_widths[2], 6, str(item['Size (GB)']), border=1, align='R')
                pdf.cell(col_widths[3], 6, str(item['Extension']), border=1, align='C')
                pdf.ln()
            
            pdf.output(output_path)
            result_text = f"""
📄 PDF Report Created Successfully!

📁 Total Items: {len(self.files_data)}
📄 File Count: {files_count}
📂 Folder Count: {folders_count}
❌ Error Count: {error_count}
💾 Total Size: {round(total_size_gb, 2)} GB
📄 PDF File: {output_filename}
📍 Location: {current_dir}
            """
            
            self.result_label.config(text=result_text)
            self.status_label.config(text="PDF report created successfully!")
            messagebox.showinfo(
                "Success", 
                f"PDF report created!\n\n"
                f"File: {output_filename}\n"
                f"Location: {current_dir}\n\n"
                f"Total {len(self.files_data)} items included in report.\n"
                f"({files_count} files, {folders_count} folders, {error_count} errors)"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while creating PDF:\n{str(e)}")
            self.status_label.config(text="PDF creation failed!")
    
    def export_to_html(self):
        if not self.selected_folder:
            messagebox.showwarning("Warning", "Please select a folder first!")
            return
            
        if not self.files_data:
            messagebox.showinfo("Info", "Please run analysis first by creating Excel report!")
            return
            
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            selected_folder_name = os.path.basename(self.selected_folder)
            drive_letter = os.path.splitdrive(self.selected_folder)[0]
            
            try:
                import shutil
                total, used, free = shutil.disk_usage(self.selected_folder)
                drive_size_gb = total // (1024**3)
                drive_name = f"{drive_letter} ({drive_size_gb}GB)"
            except:
                drive_name = drive_letter
            
            output_filename = f"{drive_name}_{selected_folder_name}_analysis_{timestamp}.html"
            current_dir = os.path.dirname(os.path.abspath(__file__))
            output_path = os.path.join(current_dir, output_filename)
            files_count = len([item for item in self.files_data if item['Type'] == 'File'])
            folders_count = len([item for item in self.files_data if item['Type'] == 'Folder'])
            error_count = len([item for item in self.files_data if item['Type'] == 'Error'])
            total_size_gb = sum(item['Size (GB)'] for item in self.files_data if item['Type'] in ['File', 'Folder'])
            df = pd.DataFrame(self.files_data)
            html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>File Size Analysis Report</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 28px;
            font-weight: 300;
        }}
        .header p {{
            margin: 10px 0 0 0;
            opacity: 0.9;
            font-size: 14px;
        }}
        .summary {{
            background-color: #f8f9fa;
            padding: 20px;
            border-bottom: 1px solid #dee2e6;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-top: 15px;
        }}
        .summary-item {{
            background: white;
            padding: 15px;
            border-radius: 6px;
            text-align: center;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .summary-item h3 {{
            margin: 0 0 5px 0;
            color: #495057;
            font-size: 14px;
        }}
        .summary-item .value {{
            font-size: 24px;
            font-weight: bold;
            color: #007bff;
        }}
        .table-container {{
            padding: 20px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            background: white;
            border-radius: 6px;
            overflow: hidden;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        th {{
            background-color: #007bff;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 10px 12px;
            border-bottom: 1px solid #dee2e6;
        }}
        tr:hover {{
            background-color: #f8f9fa;
        }}
        .folder-row {{
            background-color: #fff3cd;
        }}
        .file-row {{
            background-color: #d1ecf1;
        }}
        .error-row {{
            background-color: #f8d7da;
        }}
        .footer {{
            background-color: #343a40;
            color: white;
            text-align: center;
            padding: 15px;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📁 {drive_name} - {selected_folder_name}</h1>
            <p>Detailed File Analysis Report</p>
            <p>📊 Analysis Date: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}</p>
            <p>📍 Folder: {self.selected_folder}</p>
        </div>
        
        <div class="summary">
            <h2>📊 Analysis Summary</h2>
            <div class="summary-grid">
                <div class="summary-item">
                    <h3>Total Items</h3>
                    <div class="value">{len(self.files_data)}</div>
                </div>
                <div class="summary-item">
                    <h3>Files</h3>
                    <div class="value">{files_count}</div>
                </div>
                <div class="summary-item">
                    <h3>Folders</h3>
                    <div class="value">{folders_count}</div>
                </div>
                <div class="summary-item">
                    <h3>Errors</h3>
                    <div class="value">{error_count}</div>
                </div>
                <div class="summary-item">
                    <h3>Total Size</h3>
                    <div class="value">{round(total_size_gb, 2)} GB</div>
                </div>
            </div>
        </div>
        
        <div class="table-container">
            <h2>📋 Detailed File Analysis</h2>
            {df.to_html(classes='table table-striped', index=False, escape=False)}
        </div>
        
        <div class="footer">
            <p>Generated by File Size Analyzer | {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
            """
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            result_text = f"""
🌐 HTML Report Created Successfully!

📁 Total Items: {len(self.files_data)}
📄 File Count: {files_count}
📂 Folder Count: {folders_count}
❌ Error Count: {error_count}
💾 Total Size: {round(total_size_gb, 2)} GB
🌐 HTML File: {output_filename}
📍 Location: {current_dir}
            """
            
            self.result_label.config(text=result_text)
            self.status_label.config(text="HTML report created successfully!")
            
            messagebox.showinfo(
                "Success", 
                f"HTML report created!\n\n"
                f"File: {output_filename}\n"
                f"Location: {current_dir}\n\n"
                f"Total {len(self.files_data)} items included in report.\n"
                f"({files_count} files, {folders_count} folders, {error_count} errors)\n\n"
                f"Opening in browser..."
            )
            webbrowser.open(f'file://{output_path}')
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while creating HTML:\n{str(e)}")
            self.status_label.config(text="HTML creation failed!")
    
    def show_charts(self):
        if not self.selected_folder:
            messagebox.showwarning("Warning", "Please select a folder first!")
            return
            
        if not self.files_data:
            messagebox.showinfo("Info", "Please run analysis first by creating Excel report!")
            return
            
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            selected_folder_name = os.path.basename(self.selected_folder)
            drive_letter = os.path.splitdrive(self.selected_folder)[0]
            
            try:
                total, used, free = shutil.disk_usage(self.selected_folder)
                drive_size_gb = total // (1024**3)
                drive_name = f"{drive_letter} ({drive_size_gb}GB)"
            except:
                drive_name = drive_letter
            
            current_dir = os.path.dirname(os.path.abspath(__file__))
            
            plt.style.use('seaborn-v0_8')
            sns.set_palette("husl")
            fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(20, 6))
            fig.suptitle(f'📁 {drive_name} - {selected_folder_name}\nFile Analysis Charts', 
                        fontsize=16, fontweight='bold', y=0.95)
            
            df = pd.DataFrame(self.files_data)
            file_types = df[df['Type'] == 'File']['Extension'].value_counts()
            
            if len(file_types) > 0:
                if len(file_types) > 8:
                    top_types = file_types.head(8)
                    other_count = file_types.iloc[8:].sum()
                    top_types['Others'] = other_count
                    plot_data = top_types
                else:
                    plot_data = file_types
                
                colors = plt.cm.Set3(np.linspace(0, 1, len(plot_data)))
                wedges, texts, autotexts = ax1.pie(plot_data.values, labels=plot_data.index, 
                                                  autopct='%1.1f%%', startangle=90, colors=colors)
                ax1.set_title('📄 File Types Distribution', fontsize=14, fontweight='bold', pad=20)
                
                for autotext in autotexts:
                    autotext.set_color('white')
                    autotext.set_fontweight('bold')
            else:
                ax1.text(0.5, 0.5, 'No files found', ha='center', va='center', 
                        transform=ax1.transAxes, fontsize=12)
                ax1.set_title('📄 File Types Distribution', fontsize=14, fontweight='bold', pad=20)
            
            folder_data = df[df['Type'] == 'Folder'].sort_values('Size (GB)', ascending=False).head(10)
            
            if len(folder_data) > 0:
                bars = ax2.barh(range(len(folder_data)), folder_data['Size (GB)'], 
                               color=plt.cm.viridis(np.linspace(0, 1, len(folder_data))))
                ax2.set_yticks(range(len(folder_data)))
                ax2.set_yticklabels([name[:20] + '...' if len(name) > 20 else name 
                                   for name in folder_data['Name']])
                ax2.set_xlabel('Size (GB)', fontweight='bold')
                ax2.set_title('📂 Top 10 Largest Folders', fontsize=14, fontweight='bold', pad=20)
                
                for i, (bar, size) in enumerate(zip(bars, folder_data['Size (GB)'])):
                    ax2.text(bar.get_width() + 0.01, bar.get_y() + bar.get_height()/2, 
                            f'{size:.2f} GB', va='center', fontweight='bold')
                
                ax2.invert_yaxis()
            else:
                ax2.text(0.5, 0.5, 'No folders found', ha='center', va='center', 
                        transform=ax2.transAxes, fontsize=12)
                ax2.set_title('📂 Top 10 Largest Folders', fontsize=14, fontweight='bold', pad=20)
            
            folder_data_all = df[df['Type'] == 'Folder'].sort_values('Size (GB)', ascending=False)
            
            if len(folder_data_all) > 0:
                max_folders = min(20, len(folder_data_all))
                grid_size = int(np.ceil(np.sqrt(max_folders)))
                heatmap_data = np.zeros((grid_size, grid_size))
                folder_names = []
                
                for i in range(max_folders):
                    row = i // grid_size
                    col = i % grid_size
                    heatmap_data[row, col] = folder_data_all.iloc[i]['Size (GB)']
                    folder_names.append(folder_data_all.iloc[i]['Name'][:15] + '...' 
                                      if len(folder_data_all.iloc[i]['Name']) > 15 
                                      else folder_data_all.iloc[i]['Name'])
                
                im = ax3.imshow(heatmap_data, cmap='magma', aspect='auto')
                ax3.set_title('🔥 Folder Size Heat Map', fontsize=14, fontweight='bold', pad=20)
                cbar = plt.colorbar(im, ax=ax3, shrink=0.8)
                cbar.set_label('Size (GB)', fontweight='bold')
                
                for i in range(max_folders):
                    row = i // grid_size
                    col = i % grid_size
                    size = heatmap_data[row, col]
                    if size > 0:
                        ax3.text(col, row, f'{folder_names[i]}\n{size:.2f} GB', 
                                ha='center', va='center', fontsize=8, fontweight='bold',
                                color='white' if size > heatmap_data.max() * 0.5 else 'black')
                
                ax3.set_xticks([])
                ax3.set_yticks([])
                
            else:
                ax3.text(0.5, 0.5, 'No folders found', ha='center', va='center', 
                        transform=ax3.transAxes, fontsize=12)
                ax3.set_title('🔥 Folder Size Heat Map', fontsize=14, fontweight='bold', pad=20)
            
            plt.tight_layout()
            chart_filename = f"{drive_name}_{selected_folder_name}_charts_{timestamp}.png"
            chart_path = os.path.join(current_dir, chart_filename)
            plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
            plt.show()
            result_text = f"""
📈 Charts Generated Successfully!

📁 Total Items: {len(self.files_data)}
📄 File Types: {len(df[df['Type'] == 'File']['Extension'].unique())}
📂 Folders Analyzed: {len(df[df['Type'] == 'Folder'])}
📊 Chart File: {chart_filename}
📍 Location: {current_dir}
            """
            
            self.result_label.config(text=result_text)
            self.status_label.config(text="Charts generated and displayed successfully!")
            messagebox.showinfo(
                "Success", 
                f"Charts generated successfully!\n\n"
                f"File: {chart_filename}\n"
                f"Location: {current_dir}\n\n"
                f"Charts displayed in new window.\n"
                f"Chart saved as PNG file."
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while creating charts:\n{str(e)}")
            self.status_label.config(text="Chart creation failed!")
    
    def get_file_category(self, file_extension):
        for category, extensions in self.file_categories.items():
            if file_extension.lower() in extensions:
                return category
        return "Other"
    
    def calculate_compression_savings(self, file_size_gb, file_category):
        savings = {}
        
        if file_category in self.compression_algorithms:
            for algorithm, efficiency in self.compression_algorithms[file_category].items():
                original_size = file_size_gb
                compressed_size = original_size * (1 - efficiency)
                savings_gb = original_size - compressed_size
                savings_percentage = efficiency * 100
                savings[algorithm] = {
                    'original_size': original_size,
                    'compressed_size': compressed_size,
                    'savings_gb': savings_gb,
                    'savings_percentage': savings_percentage
                }
        return savings
    
    def show_optimization(self):
        if not self.files_data:
            messagebox.showwarning("Warning", "Please analyze a folder first!")
            return
        
        try:
            self.progress.start()
            self.status_label.config(text="Analyzing optimization opportunities...")
            opt_window = tk.Toplevel(self.root)
            opt_window.title("💾 Size Optimization Analysis")
            opt_window.geometry("1000x700")
            opt_window.configure(bg=self.colors['bg'])
            opt_window.resizable(True, True)
            opt_window.transient(self.root)
            opt_window.grab_set()
            opt_window.update_idletasks()
            x = (opt_window.winfo_screenwidth() // 2) - (1000 // 2)
            y = (opt_window.winfo_screenheight() // 2) - (700 // 2)
            opt_window.geometry(f"1000x700+{x}+{y}")
            title_label = tk.Label(
                opt_window,
                text="💾 Size Optimization Analysis",
                font=("Segoe UI", 16, "bold"),
                bg=self.colors['bg'],
                fg=self.colors['fg']
            )
            title_label.pack(pady=(20, 10))
            desc_label = tk.Label(
                opt_window,
                text="Potential space savings with different compression algorithms",
                font=("Segoe UI", 10),
                bg=self.colors['bg'],
                fg=self.colors['text_secondary']
            )
            desc_label.pack(pady=(0, 20))
            main_frame = tk.Frame(opt_window, bg=self.colors['bg'])
            main_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))
            canvas = tk.Canvas(main_frame, bg=self.colors['bg'], highlightthickness=0)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg=self.colors['bg'])
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            summary_frame = tk.Frame(scrollable_frame, bg=self.colors['secondary_bg'], relief='flat', bd=1)
            summary_frame.pack(fill='x', pady=(0, 20))
            summary_label = tk.Label(
                summary_frame,
                text="📊 Optimization Summary",
                font=("Segoe UI", 12, "bold"),
                bg=self.colors['secondary_bg'],
                fg=self.colors['fg']
            )
            summary_label.pack(pady=(10, 5))
            total_original_size = 0
            total_folder_size = 0
            total_potential_savings = {}
            files_list = [item for item in self.files_data if item['Type'] == 'File']
            folders_list = [item for item in self.files_data if item['Type'] == 'Folder']
            
            for item in files_list:
                file_size = item['Size (GB)']
                total_original_size += file_size
                
                file_category = self.get_file_category(item['Extension'])
                savings = self.calculate_compression_savings(file_size, file_category)
                
                for algorithm, data in savings.items():
                    if algorithm not in total_potential_savings:
                        total_potential_savings[algorithm] = 0
                    total_potential_savings[algorithm] += data['savings_gb']
            
            for item in folders_list:
                total_folder_size += item['Size (GB)']
            
            summary_text = f"📁 Total Items: {len(self.files_data)}\n"
            summary_text += f"📄 Files: {len(files_list)} ({total_original_size:.2f} GB)\n"
            summary_text += f"📂 Folders: {len(folders_list)} ({total_folder_size:.2f} GB)\n"
            summary_text += f"💾 Total Size: {total_original_size + total_folder_size:.2f} GB\n\n"
            summary_text += "🔥 Top 5 Potential Savings:\n"
            sorted_algorithms = sorted(total_potential_savings.items(), key=lambda x: x[1], reverse=True)
            
            for i, (algorithm, total_savings) in enumerate(sorted_algorithms[:5]):
                percentage = (total_savings / total_original_size * 100) if total_original_size > 0 else 0
                summary_text += f"{i+1}. {algorithm}: {total_savings:.2f} GB ({percentage:.1f}%)\n"
            
            summary_content = tk.Label(
                summary_frame,
                text=summary_text,
                font=("Segoe UI", 10),
                bg=self.colors['secondary_bg'],
                fg=self.colors['fg'],
                justify='left'
            )
            summary_content.pack(pady=(0, 10), padx=10)
            
            detail_label = tk.Label(
                scrollable_frame,
                text="📋 Detailed Item Analysis (Files & Folders)",
                font=("Segoe UI", 12, "bold"),
                bg=self.colors['bg'],
                fg=self.colors['fg']
            )
            detail_label.pack(pady=(0, 10), anchor='w')
            
            for item in self.files_data:
                item_frame = tk.Frame(scrollable_frame, bg=self.colors['secondary_bg'], relief='flat', bd=1)
                item_frame.pack(fill='x', pady=(0, 10))
                item_header = tk.Frame(item_frame, bg=self.colors['secondary_bg'])
                item_header.pack(fill='x', padx=10, pady=(10, 5))
                icon = "📄" if item['Type'] == 'File' else "📂"
                item_name_label = tk.Label(
                    item_header,
                    text=f"{icon} {item['Name']}",
                    font=("Segoe UI", 10, "bold"),
                    bg=self.colors['secondary_bg'],
                    fg=self.colors['fg']
                )
                item_name_label.pack(side='left')
                item_size_label = tk.Label(
                    item_header,
                    text=f"Size: {item['Size (GB)']:.2f} GB",
                    font=("Segoe UI", 9),
                    bg=self.colors['secondary_bg'],
                    fg=self.colors['text_secondary']
                )
                item_size_label.pack(side='right')
                
                if item['Type'] == 'File':
                    file_category = self.get_file_category(item['Extension'])
                    savings = self.calculate_compression_savings(item['Size (GB)'], file_category)
                    
                    if savings:
                        savings_frame = tk.Frame(item_frame, bg=self.colors['secondary_bg'])
                        savings_frame.pack(fill='x', padx=10, pady=(0, 10))
                        savings_label = tk.Label(
                            savings_frame,
                            text=f"Category: {file_category}",
                            font=("Segoe UI", 9),
                            bg=self.colors['secondary_bg'],
                            fg=self.colors['text_secondary']
                        )
                        savings_label.pack(anchor='w', pady=(0, 5))
                        algo_frame = tk.Frame(savings_frame, bg=self.colors['secondary_bg'])
                        algo_frame.pack(fill='x')
                        
                        for i, (algorithm, data) in enumerate(savings.items()):
                            algo_button = tk.Button(
                                algo_frame,
                                text=f"{algorithm}\n{data['savings_gb']:.2f} GB saved ({data['savings_percentage']:.0f}%)",
                                font=("Segoe UI", 8),
                                bg=self.colors['accent'],
                                fg=self.colors['fg'],
                                relief='flat',
                                padx=8,
                                pady=4,
                                cursor='hand2',
                                activebackground='#005a9e',
                                activeforeground=self.colors['fg']
                            )
                            algo_button.grid(row=i//2, column=i%2, padx=(0, 10), pady=(0, 5), sticky='ew')
                            algo_frame.grid_columnconfigure(0, weight=1)
                            algo_frame.grid_columnconfigure(1, weight=1)
                    else:
                        no_savings_label = tk.Label(
                            item_frame,
                            text="No optimization algorithms available for this file type",
                            font=("Segoe UI", 9),
                            bg=self.colors['secondary_bg'],
                            fg=self.colors['text_secondary']
                        )
                        no_savings_label.pack(pady=(0, 10), padx=10)
                else:
                    folder_info_label = tk.Label(
                        item_frame,
                        text="📂 Folder - Contains multiple files that can be optimized individually",
                        font=("Segoe UI", 9),
                        bg=self.colors['secondary_bg'],
                        fg=self.colors['text_secondary']
                    )
                    folder_info_label.pack(pady=(0, 10), padx=10)
            
            export_frame = tk.Frame(opt_window, bg=self.colors['bg'])
            export_frame.pack(fill='x', padx=20, pady=(0, 20))
            export_btn = tk.Button(
                export_frame,
                text="📊 Export Optimization Report",
                command=lambda: self.export_optimization_report(opt_window),
                font=("Segoe UI", 11, "bold"),
                bg=self.colors['success'],
                fg=self.colors['fg'],
                relief='flat',
                padx=15,
                pady=8,
                cursor='hand2',
                activebackground='#1e7e34',
                activeforeground=self.colors['fg']
            )
            export_btn.pack()
            self.progress.stop()
            self.status_label.config(text="Optimization analysis completed!")
            
        except Exception as e:
            self.progress.stop()
            messagebox.showerror("Error", f"An error occurred during optimization analysis:\n{str(e)}")
            self.status_label.config(text="Optimization analysis failed!")
    
    def export_optimization_report(self, parent_window):
        try:
            if not self.files_data:
                messagebox.showwarning("Warning", "No data to export!")
                return
            optimization_data = []
            
            for item in self.files_data:
                if item['Type'] == 'File':
                    file_category = self.get_file_category(item['Extension'])
                    savings = self.calculate_compression_savings(item['Size (GB)'], file_category)
                    
                    if savings:
                        for algorithm, data in savings.items():
                            optimization_data.append({
                                'File Name': item['Name'],
                                'File Extension': item['Extension'],
                                'File Category': file_category,
                                'Original Size (GB)': item['Size (GB)'],
                                'Algorithm': algorithm,
                                'Compressed Size (GB)': data['compressed_size'],
                                'Savings (GB)': data['savings_gb'],
                                'Savings (%)': data['savings_percentage']
                            })
                    else:
                        optimization_data.append({
                            'File Name': item['Name'],
                            'File Extension': item['Extension'],
                            'File Category': file_category,
                            'Original Size (GB)': item['Size (GB)'],
                            'Algorithm': 'No optimization available',
                            'Compressed Size (GB)': item['Size (GB)'],
                            'Savings (GB)': 0,
                            'Savings (%)': 0
                        })
            
            df = pd.DataFrame(optimization_data)
            current_dir = os.path.dirname(os.path.abspath(__file__))
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            selected_folder_name = os.path.basename(self.selected_folder)
            drive_letter = os.path.splitdrive(self.selected_folder)[0]
            
            try:
                import shutil
                total, used, free = shutil.disk_usage(self.selected_folder)
                drive_size_gb = total // (1024**3)
                drive_name = f"{drive_letter} ({drive_size_gb}GB)"
            except:
                drive_name = drive_letter
            
            filename = f"{drive_name}_{selected_folder_name}_optimization_{timestamp}.xlsx"
            filepath = os.path.join(current_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Optimization Analysis', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Optimization Analysis']
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                title = f"💾 Size Optimization Report - {selected_folder_name}"
                worksheet.insert_rows(1, 2)
                worksheet['A1'] = title
                worksheet.merge_cells('A1:H1')
                
                from openpyxl.styles import Font, Alignment, PatternFill
                title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
                title_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
                title_alignment = Alignment(horizontal='center', vertical='center')
                
                worksheet['A1'].font = title_font
                worksheet['A1'].fill = title_fill
                worksheet['A1'].alignment = title_alignment
                
                header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
                
                for cell in worksheet[3]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            messagebox.showinfo(
                "Success",
                f"Optimization report exported successfully!\n\n"
                f"File: {filename}\n"
                f"Location: {current_dir}"
            )
            parent_window.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while exporting optimization report:\n{str(e)}")
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = FileSizeAnalyzer()
    app.run() 