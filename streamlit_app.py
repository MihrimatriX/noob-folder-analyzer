import streamlit as st
import pandas as pd
import os
from pathlib import Path
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import base64
from io import BytesIO
import zipfile
import json
import platform

def normalize_windows_path(path):
    """Normalize Windows path format for Docker environment"""
    if not path:
        return path
    
    # Docker container'da Windows drive'larına erişim için path'i dönüştür
    if platform.system() == "Linux":  # Docker container'da Linux çalışır
        # Önce mevcut mount'ları kontrol et
        available_mounts = []
        for drive in 'cdefghijklmnopqrstuvwxyz':
            mount_path = f'/host/{drive}'
            if os.path.exists(mount_path):
                available_mounts.append(drive.upper())
        
        # Windows path formatını Docker path formatına çevir
        if path.startswith('C:') or path.startswith('c:'):
            if 'C' in available_mounts:
                return '/host/c/' + path[3:].replace('\\', '/')
            else:
                return None  # C: sürücüsü mevcut değil
        elif path.startswith('D:') or path.startswith('d:'):
            if 'D' in available_mounts:
                return '/host/d/' + path[3:].replace('\\', '/')
            else:
                return None  # D: sürücüsü mevcut değil
        elif path.startswith('E:') or path.startswith('e:'):
            if 'E' in available_mounts:
                return '/host/e/' + path[3:].replace('\\', '/')
            else:
                return None  # E: sürücüsü mevcut değil
        elif path.startswith('F:') or path.startswith('f:'):
            if 'F' in available_mounts:
                return '/host/f/' + path[3:].replace('\\', '/')
            else:
                return None  # F: sürücüsü mevcut değil
        elif path.startswith('G:') or path.startswith('g:'):
            if 'G' in available_mounts:
                return '/host/g/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('H:') or path.startswith('h:'):
            if 'H' in available_mounts:
                return '/host/h/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('I:') or path.startswith('i:'):
            if 'I' in available_mounts:
                return '/host/i/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('J:') or path.startswith('j:'):
            if 'J' in available_mounts:
                return '/host/j/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('K:') or path.startswith('k:'):
            if 'K' in available_mounts:
                return '/host/k/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('L:') or path.startswith('l:'):
            if 'L' in available_mounts:
                return '/host/l/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('M:') or path.startswith('m:'):
            if 'M' in available_mounts:
                return '/host/m/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('N:') or path.startswith('n:'):
            if 'N' in available_mounts:
                return '/host/n/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('O:') or path.startswith('o:'):
            if 'O' in available_mounts:
                return '/host/o/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('P:') or path.startswith('p:'):
            if 'P' in available_mounts:
                return '/host/p/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('Q:') or path.startswith('q:'):
            if 'Q' in available_mounts:
                return '/host/q/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('R:') or path.startswith('r:'):
            if 'R' in available_mounts:
                return '/host/r/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('S:') or path.startswith('s:'):
            if 'S' in available_mounts:
                return '/host/s/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('T:') or path.startswith('t:'):
            if 'T' in available_mounts:
                return '/host/t/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('U:') or path.startswith('u:'):
            if 'U' in available_mounts:
                return '/host/u/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('V:') or path.startswith('v:'):
            if 'V' in available_mounts:
                return '/host/v/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('W:') or path.startswith('w:'):
            if 'W' in available_mounts:
                return '/host/w/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('X:') or path.startswith('x:'):
            if 'X' in available_mounts:
                return '/host/x/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('Y:') or path.startswith('y:'):
            if 'Y' in available_mounts:
                return '/host/y/' + path[3:].replace('\\', '/')
            else:
                return None
        elif path.startswith('Z:') or path.startswith('z:'):
            if 'Z' in available_mounts:
                return '/host/z/' + path[3:].replace('\\', '/')
            else:
                return None
    
    elif platform.system() == "Windows":
        # Native Windows ortamında normal path dönüşümü
        if path.endswith('\\') and len(path) == 2:
            normalized = path.rstrip('\\')
            if len(normalized) == 1:
                return normalized + ":\\"
            return normalized
        path = path.replace('\\\\', '\\')
        if len(path) >= 2 and path[1] == ':' and len(path) == 2:
            path = path + "\\"
    
    return path.strip()

def get_folder_name(path):
    """Extract folder name from full path"""
    if not path:
        return path
    return os.path.basename(path) if os.path.basename(path) else os.path.dirname(path).split(os.sep)[-1]

st.set_page_config(
    page_title="File Size Analyzer",
    page_icon="📁",
    layout="wide",
    initial_sidebar_state="collapsed"
)
st.markdown("""
<style>
    @import url('https://cdn.jsdelivr.net/npm/tailwindcss@3.3.0/dist/tailwind.min.css');
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600&display=swap');

    :root {
        --primary-gradient: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        --secondary-gradient: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        --accent-gradient: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        --success-gradient: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%);
        --warning-gradient: linear-gradient(135deg, #fa709a 0%, #fee140 100%);
        --dark-bg: #0f172a;
        --dark-surface: #1e293b;
        --dark-surface-2: #334155;
        --text-primary: #f8fafc;
        --text-secondary: #cbd5e1;
        --text-muted: #64748b;
        --border-color: rgba(148, 163, 184, 0.1);
        --shadow-sm: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
        --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
        --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
        --shadow-xl: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04);
        --shadow-2xl: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
    }
    
    * {
        box-sizing: border-box;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    }
    
    .main .block-container {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #334155 100%);
        min-height: 100vh;
        padding: 1rem 0.5rem;
        max-width: 800px !important;
        margin: 0 auto !important;
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        position: relative;
        overflow-x: hidden;
    }
    
    .stMainBlockContainer {
        max-width: 800px !important;
        margin: 0 auto !important;
    }
    
    .main .block-container::before {
        content: '';
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        background-image: 
            radial-gradient(circle at 20% 80%, rgba(120, 119, 198, 0.1) 0%, transparent 50%),
            radial-gradient(circle at 80% 20%, rgba(255, 119, 198, 0.1) 0%, transparent 50%),
            radial-gradient(circle at 40% 40%, rgba(120, 219, 255, 0.1) 0%, transparent 50%);
        pointer-events: none;
        z-index: -1;
    }
    
    .main-header {
        font-size: 2.5rem;
        font-weight: 900;
        text-align: center;
        margin-bottom: 0.75rem;
        background: var(--primary-gradient);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        letter-spacing: -0.025em;
        position: relative;
        animation: titleGlow 3s ease-in-out infinite alternate;
    }
    
    @keyframes titleGlow {
        0% { filter: drop-shadow(0 0 20px rgba(102, 126, 234, 0.3)); }
        100% { filter: drop-shadow(0 0 30px rgba(118, 75, 162, 0.5)); }
    }
    
    .sub-header {
        font-size: 1rem;
        color: var(--text-secondary);
        text-align: center;
        margin-bottom: 1rem;
        line-height: 1.6;
        font-weight: 400;
        opacity: 0.9;
    }
    
    .main-content {
        background: rgba(30, 41, 59, 0.8);
        backdrop-filter: blur(20px) saturate(180%);
        border-radius: 24px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        box-shadow: var(--shadow-2xl);
        border: 1px solid var(--border-color);
        max-width: 750px !important;
        margin-left: auto !important;
        margin-right: auto !important;
        width: 100% !important;
        position: relative;
        overflow: hidden;
    }
    
    .main-content::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 1px;
        background: var(--primary-gradient);
        opacity: 0.5;
    }
    
    .metric-card {
        background: rgba(51, 65, 85, 0.6);
        backdrop-filter: blur(10px);
        border-radius: 16px;
        padding: 1rem;
        margin: 0.25rem 0;
        border: 1px solid var(--border-color);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        overflow: hidden;
    }
    
    .metric-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        background: var(--primary-gradient);
        opacity: 0;
        transition: opacity 0.3s ease;
        z-index: -1;
    }
    
    .metric-card:hover {
        transform: translateY(-4px) scale(1.02);
        box-shadow: var(--shadow-xl);
        border-color: rgba(102, 126, 234, 0.3);
    }
    
    .metric-card:hover::before {
        opacity: 0.1;
    }
    
    .filter-form {
        background: rgba(30, 41, 59, 0.9);
        backdrop-filter: blur(20px) saturate(180%);
        border-radius: 20px;
        padding: 1.5rem;
        box-shadow: var(--shadow-xl);
        border: 1px solid var(--border-color);
        max-width: 700px !important;
        margin: 0 auto !important;
        width: 100% !important;
        position: relative;
    }
    
    .filter-form::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 2px;
        background: var(--secondary-gradient);
        border-radius: 20px 20px 0 0;
    }
    
    .folder-info {
        background: rgba(51, 65, 85, 0.7);
        backdrop-filter: blur(15px);
        border-radius: 16px;
        padding: 1rem;
        color: var(--text-secondary);
        font-size: 0.875rem;
        margin: 0.5rem auto !important;
        max-width: 750px !important;
        width: 100% !important;
        border: 1px solid var(--border-color);
        box-shadow: var(--shadow-lg);
    }
    
    .stTextInput > div > div > input {
        background: rgba(51, 65, 85, 0.8) !important;
        border: 1px solid var(--border-color) !important;
        color: var(--text-primary) !important;
        border-radius: 12px !important;
        padding: 0.75rem 1rem !important;
        font-size: 0.875rem !important;
        transition: all 0.3s ease !important;
        backdrop-filter: blur(10px) !important;
    }
    
    .stTextInput > div > div > input:focus {
        border-color: #667eea !important;
        box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1) !important;
        outline: none !important;
    }
    
    .stSelectbox > div > div > div {
        background: rgba(51, 65, 85, 0.8) !important;
        border: 1px solid var(--border-color) !important;
        color: var(--text-primary) !important;
        border-radius: 12px !important;
        font-size: 0.875rem !important;
        backdrop-filter: blur(10px) !important;
    }
    
    .stMultiselect > div > div > div {
        background: rgba(51, 65, 85, 0.8) !important;
        border: 1px solid var(--border-color) !important;
        color: var(--text-primary) !important;
        border-radius: 12px !important;
        font-size: 0.875rem !important;
        backdrop-filter: blur(10px) !important;
    }
    
    .stNumberInput > div > div > input {
        background: rgba(51, 65, 85, 0.8) !important;
        border: 1px solid var(--border-color) !important;
        color: var(--text-primary) !important;
        border-radius: 12px !important;
        padding: 0.75rem 1rem !important;
        font-size: 0.875rem !important;
        backdrop-filter: blur(10px) !important;
    }
    
    .stDateInput > div > div > input {
        background: rgba(51, 65, 85, 0.8) !important;
        border: 1px solid var(--border-color) !important;
        color: var(--text-primary) !important;
        border-radius: 12px !important;
        padding: 0.75rem 1rem !important;
        font-size: 0.875rem !important;
        backdrop-filter: blur(10px) !important;
    }
    
    .stCheckbox > div > div {
        background: rgba(51, 65, 85, 0.8) !important;
        border: 1px solid var(--border-color) !important;
        border-radius: 8px !important;
        backdrop-filter: blur(10px) !important;
    }
    
    .stCheckbox > label {
        color: var(--text-secondary) !important;
        font-size: 0.875rem !important;
        font-weight: 500 !important;
    }
    
    .stButton > button {
        background: var(--primary-gradient) !important;
        color: white !important;
        font-weight: 600 !important;
        padding: 0.75rem 1.5rem !important;
        border-radius: 12px !important;
        border: none !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        font-size: 0.875rem !important;
        position: relative !important;
        overflow: hidden !important;
        box-shadow: var(--shadow-md) !important;
    }
    
    .stButton > button::before {
        content: '';
        position: absolute;
        top: 0;
        left: -100%;
        width: 100%;
        height: 100%;
        background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.2), transparent);
        transition: left 0.5s;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: var(--shadow-xl) !important;
    }
    
    .stButton > button:hover::before {
        left: 100%;
    }
    
    .stTabs > div > div > div > div {
        background: rgba(30, 41, 59, 0.8) !important;
        border: 1px solid var(--border-color) !important;
        border-radius: 16px 16px 0 0 !important;
        backdrop-filter: blur(15px) !important;
    }
    
    .stTabs > div > div > div > div > button {
        color: var(--text-secondary) !important;
        font-weight: 500 !important;
        font-size: 0.875rem !important;
        padding: 0.75rem 1rem !important;
        border-radius: 12px 12px 0 0 !important;
        transition: all 0.3s ease !important;
        background: transparent !important;
        border: none !important;
    }
    
    .stTabs > div > div > div > div > button:hover {
        color: var(--text-primary) !important;
        background: rgba(102, 126, 234, 0.1) !important;
    }
    
    .stTabs > div > div > div > div > button[aria-selected="true"] {
        color: white !important;
        background: var(--primary-gradient) !important;
        box-shadow: var(--shadow-md) !important;
    }
    
    .stDataFrame > div > div {
        background: rgba(30, 41, 59, 0.8) !important;
        border: 1px solid var(--border-color) !important;
        border-radius: 16px !important;
        backdrop-filter: blur(15px) !important;
        overflow: hidden !important;
    }
    
    .stSubheader {
        font-size: 1.25rem !important;
        font-weight: 700 !important;
        color: var(--text-primary) !important;
        margin-bottom: 1rem !important;
        background: var(--accent-gradient) !important;
        -webkit-background-clip: text !important;
        -webkit-text-fill-color: transparent !important;
        background-clip: text !important;
    }
    
    .stMetric {
        background: rgba(51, 65, 85, 0.6) !important;
        backdrop-filter: blur(10px) !important;
        border-radius: 16px !important;
        padding: 1rem !important;
        border: 1px solid var(--border-color) !important;
        box-shadow: var(--shadow-md) !important;
        transition: all 0.3s ease !important;
    }
    
    .stMetric:hover {
        transform: translateY(-2px) !important;
        box-shadow: var(--shadow-lg) !important;
        border-color: rgba(102, 126, 234, 0.3) !important;
    }
    
    .stSpinner > div {
        color: #667eea !important;
        filter: drop-shadow(0 0 10px rgba(102, 126, 234, 0.5)) !important;
    }
    
    .stAlert {
        border-radius: 16px !important;
        border: none !important;
        font-size: 0.875rem !important;
        backdrop-filter: blur(15px) !important;
        box-shadow: var(--shadow-lg) !important;
    }
    
    @media (max-width: 640px) {
        .main .block-container {
            padding: 0.5rem 0.25rem !important;
            max-width: 100% !important;
        }
        
        .main-content {
            padding: 1rem !important;
            max-width: 100% !important;
            border-radius: 20px !important;
        }
        
        .filter-form {
            padding: 1rem !important;
            max-width: 100% !important;
            border-radius: 16px !important;
        }
        
        .main-header {
            font-size: 2rem !important;
        }
        
        .sub-header {
            font-size: 0.875rem !important;
        }
        
        .stTabs > div > div > div > div > button {
            font-size: 0.75rem !important;
            padding: 0.5rem 0.75rem !important;
        }
        
        .stMetric {
            padding: 0.75rem !important;
        }
    }
    
    @media (min-width: 641px) and (max-width: 1024px) {
        .main .block-container {
            padding: 1rem 0.5rem !important;
            max-width: 95% !important;
        }
        
        .main-content {
            padding: 1.25rem !important;
            max-width: 95% !important;
        }
        
        .filter-form {
            padding: 1.25rem !important;
            max-width: 95% !important;
        }
        
        .main-header {
            font-size: 2.25rem !important;
        }
    }
    
    @media (min-width: 1025px) {
        .main .block-container {
            padding: 1.5rem !important;
            max-width: 800px !important;
        }
        
        .main-content {
            padding: 2rem !important;
            max-width: 750px !important;
        }
        
        .filter-form {
            padding: 2rem !important;
            max-width: 700px !important;
        }
    }
    
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(30px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    @keyframes slideInLeft {
        from {
            opacity: 0;
            transform: translateX(-30px);
        }
        to {
            opacity: 1;
            transform: translateX(0);
        }
    }
    
    @keyframes pulse {
        0%, 100% {
            opacity: 1;
        }
        50% {
            opacity: 0.7;
        }
    }
    
    .main-content {
        animation: fadeInUp 0.6s cubic-bezier(0.4, 0, 0.2, 1);
    }
    
    .filter-form {
        animation: slideInLeft 0.6s cubic-bezier(0.4, 0, 0.2, 1) 0.2s both;
    }
    
    .metric-card {
        animation: fadeInUp 0.6s cubic-bezier(0.4, 0, 0.2, 1) 0.4s both;
    }
    
    ::-webkit-scrollbar {
        width: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: rgba(30, 41, 59, 0.5);
        border-radius: 4px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: var(--primary-gradient);
        border-radius: 4px;
        transition: all 0.3s ease;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: var(--secondary-gradient);
    }
    
    .loading {
        animation: pulse 2s cubic-bezier(0.4, 0, 0.6, 1) infinite;
    }
    
    *:focus {
        outline: 2px solid #667eea;
        outline-offset: 2px;
    }
    
    ::selection {
        background: rgba(102, 126, 234, 0.3);
        color: var(--text-primary);
    }
    
    * {
    * {
        scrollbar-width: thin;
        scrollbar-color: #667eea rgba(30, 41, 59, 0.5);
    }
</style>
""", unsafe_allow_html=True)

class FileSizeAnalyzerWeb:
    def __init__(self):
        self.file_categories = {
            'Documents': ['.pdf', '.doc', '.docx', '.txt', '.rtf', '.odt', '.pages'],
            'Images': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.svg', '.webp'],
            'Videos': ['.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.webm', '.m4v'],
            'Audio': ['.mp3', '.wav', '.flac', '.aac', '.ogg', '.wma', '.m4a'],
            'Archives': ['.zip', '.rar', '.7z', '.tar', '.gz', '.bz2', '.xz'],
            'Code': ['.py', '.js', '.html', '.css', '.java', '.cpp', '.c', '.php', '.sql'],
            'Spreadsheets': ['.xlsx', '.xls', '.csv', '.ods', '.numbers'],
            'Presentations': ['.pptx', '.ppt', '.odp', '.key']
        }
        
        self.compression_algorithms = {
            'Images': {
                'JPEG Quality 85%': 0.3,
                'WebP Conversion': 0.5,
                'PNG Optimization': 0.2,
                'AVIF Conversion': 0.6
            },
            'Videos': {
                'H.264 Compression': 0.4,
                'H.265/HEVC': 0.6,
                'AV1 Codec': 0.7,
                'WebM Conversion': 0.5
            },
            'Documents': {
                'PDF Compression': 0.3,
                'ZIP Archive': 0.4,
                '7-Zip Archive': 0.5
            },
            'Audio': {
                'MP3 128kbps': 0.6,
                'AAC 128kbps': 0.7,
                'OGG Vorbis': 0.5
            },
            'Archives': {
                '7-Zip LZMA2': 0.3,
                'ZIP Deflate': 0.2,
                'RAR5': 0.25
            },
            'Code': {
                'Minification': 0.4,
                'Gzip Compression': 0.3,
                'Brotli Compression': 0.35
            }
        }
    
    def get_folder_size(self, folder_path):
        total_size = 0
        try:
            for item in Path(folder_path).rglob('*'):
                if item.is_file():
                    total_size += item.stat().st_size
        except PermissionError:
            pass
        return total_size
    
    def analyze_folder_contents(self, folder_path, file_type_filter=None, size_filter=None, date_filter=None, search_filter=None):
        files_data = []
        
        try:
            selected_path = Path(folder_path)
            
            # Yol doğrulama
            if not selected_path.exists():
                return [{
                    'Name': f'Hata: Klasör bulunamadı - {folder_path}',
                    'Type': 'Error',
                    'Size (GB)': 0,
                    'Extension': '❌',
                    'Full Path': folder_path,
                    'Category': 'Error'
                }]
            
            if not selected_path.is_dir():
                return [{
                    'Name': f'Hata: Bu bir klasör değil - {folder_path}',
                    'Type': 'Error',
                    'Size (GB)': 0,
                    'Extension': '❌',
                    'Full Path': folder_path,
                    'Category': 'Error'
                }]
            
            system_folders = ['$RECYCLE.BIN', 'System Volume Information', 'RECYCLER', 'Thumbs.db']
            
            folders = []
            files = []
            
            try:
                for item in selected_path.iterdir():
                    if item.is_dir():
                        if item.name not in system_folders:
                            if search_filter and search_filter.lower() not in item.name.lower():
                                continue
                            folders.append(item)
                    elif item.is_file():
                        if item.name not in system_folders:
                            if file_type_filter:
                                file_extension = item.suffix.lower()
                                if file_extension not in file_type_filter:
                                    continue
                            
                            if size_filter:
                                size_bytes = item.stat().st_size
                                size_gb = size_bytes / (1024 * 1024 * 1024)
                                min_size, max_size = size_filter
                                if not (min_size <= size_gb <= max_size):
                                    continue
                            
                            if date_filter:
                                from datetime import datetime
                                file_creation_time = datetime.fromtimestamp(item.stat().st_ctime)
                                start_date, end_date = date_filter
                                if not (start_date <= file_creation_time <= end_date):
                                    continue
                            
                            if search_filter and search_filter.lower() not in item.name.lower():
                                continue
                            
                            files.append(item)
            except PermissionError:
                return [{
                    'Name': f'Hata: Klasöre erişim izni yok - {folder_path}',
                    'Type': 'Error',
                    'Size (GB)': 0,
                    'Extension': '❌',
                    'Full Path': folder_path,
                    'Category': 'Error'
                }]
            
            for folder in folders:
                folder_size = self.get_folder_size(folder)
                folder_size_mb = folder_size / (1024 * 1024)
                folder_size_gb = folder_size_mb / 1024
                files_data.append({
                    'Name': folder.name,
                    'Type': 'Folder',
                    'Size (GB)': round(folder_size_gb, 2),
                    'Extension': '📁',
                    'Full Path': str(folder),
                    'Category': 'Folder'
                })
            
            for file in files:
                size_bytes = file.stat().st_size
                size_mb = size_bytes / (1024 * 1024)
                size_gb = size_mb / 1024
                
                category = self.get_file_category(file.suffix.lower())
                
                files_data.append({
                    'Name': file.name,
                    'Type': 'File',
                    'Size (GB)': round(size_gb, 2),
                    'Extension': file.suffix.lower(),
                    'Full Path': str(file),
                    'Category': category
                })
            
            return files_data
            
        except Exception as e:
            return [{
                'Name': f'Error: {str(e)}',
                'Type': 'Error',
                'Size (GB)': 0,
                'Extension': '❌',
                'Full Path': folder_path,
                'Category': 'Error'
            }]
    
    def get_file_category(self, file_extension):
        for category, extensions in self.file_categories.items():
            if file_extension.lower() in extensions:
                return category
        return "Other"
    
    def calculate_compression_savings(self, file_size_gb, file_category):
        savings = {}
        
        if file_category in self.compression_algorithms:
            for algorithm, efficiency in self.compression_algorithms[file_category].items():
                original_size = file_size_gb
                compressed_size = original_size * (1 - efficiency)
                savings_gb = original_size - compressed_size
                savings_percentage = efficiency * 100
                
                savings[algorithm] = {
                    'original_size': original_size,
                    'compressed_size': compressed_size,
                    'savings_gb': savings_gb,
                    'savings_percentage': savings_percentage
                }
        
        return savings

def main():
    st.markdown('<h1 class="main-header">📁 File Size Analyzer</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Advanced file and folder analysis with intelligent optimization recommendations</p>', unsafe_allow_html=True)
    
    # Help section
    with st.expander("❓ How to use?"):
        st.markdown("""
        ### 📋 User Guide
        
        1. **📂 Folder Selection**: Enter the path of the folder you want to analyze
           - **Docker Environment**: Use Docker path format
             - `C:\\Users\\YourName\\Documents` → `/host/c/Users/YourName/Documents`
             - `E:\\Videos` → `/host/e/Videos`
             - `D:\\Photos` → `/host/d/Photos`
           - **Native Windows**: `C:\\Users\\YourName\\Documents` or `C:/Users/YourName/Documents`
           - **Mac/Linux**: `/Users/YourName/Documents`
        
        2. **🔍 Filters**: Optionally apply filters
           - File types
           - Size range
           - Date range
           - Search term
        
        3. **🚀 Analysis**: Click "Analyze Folder" button
        
        4. **📊 Results**: View analysis results in different tabs
        
        ### 💡 Tips
        - **Docker Users**: Use `/host/` prefix for Windows drives (e.g., `/host/c/Users/YourName/Documents`)
        - **Native Users**: Use normal Windows paths (e.g., `C:\\Users\\[Username]\\Documents`)
        - Make sure to enter the complete folder path
        - Avoid system folders (Windows, System32, etc.)
        - Analysis may take some time for large folders
        - Run as administrator for folders without access permission
        - Use quick option buttons to quickly access common folders
        - **Docker**: All Windows drives (C:, D:, E:, etc.) are automatically mounted
        """)
    
    with st.container():
        st.markdown("""
        <div class="bg-slate-800/95 backdrop-blur-lg rounded-xl p-4 md:p-6 mb-4 shadow-2xl border border-slate-600/20 max-w-4xl mx-auto w-full">
            <h3 class="mb-3 text-slate-100 font-semibold text-lg md:text-xl">⚙️ Analysis Controls</h3>
        </div>
        """, unsafe_allow_html=True)
        
        st.subheader("📂 Folder Selection")
        
        # Folder selection method
        selection_method = st.radio(
            "Folder selection method:",
            ["📝 Manual path input", "📁 Folder picker"],
            horizontal=True,
            help="Choose how you want to enter the folder path"
        )
        
        if selection_method == "📝 Manual path input":
            # Use suggested path if available, otherwise use text input
            if hasattr(st.session_state, 'suggested_path') and st.session_state.suggested_path:
                folder_path = st.text_input(
                    "Enter folder path:", 
                    value=st.session_state.suggested_path,
                    placeholder="C:\\Users\\YourName\\Documents",
                    help="Enter the full path of the folder to analyze"
                )
            else:
                folder_path = st.text_input(
                    "Enter folder path:", 
                    placeholder="C:\\Users\\YourName\\Documents",
                    help="Enter the full path of the folder to analyze"
                )
            
            # Quick folder options
            st.markdown("**🚀 Quick options:**")
            
            # Mevcut sürücüleri kontrol et
            available_drives = []
            if platform.system() == "Linux":  # Docker environment
                for drive in 'cdefghijklmnopqrstuvwxyz':
                    mount_path = f'/host/{drive}'
                    if os.path.exists(mount_path):
                        available_drives.append(drive.upper())
            else:  # Native Windows
                import string
                for letter in string.ascii_uppercase:
                    drive_path = f"{letter}:\\"
                    if os.path.exists(drive_path):
                        available_drives.append(letter)
            
            # Mevcut sürücüleri göster
            if available_drives:
                st.info(f"💡 Mevcut sürücüler: {', '.join(available_drives)}")
                
                # İlk 3 mevcut sürücüyü buton olarak göster
                col1, col2, col3 = st.columns(3)
                
                for i, drive in enumerate(available_drives[:3]):
                    with [col1, col2, col3][i]:
                        if st.button(f"💾 {drive}:", use_container_width=True):
                            if platform.system() == "Linux":  # Docker environment
                                drive_path = f"/host/{drive.lower()}/"
                            else:  # Native Windows
                                drive_path = f"{drive}:\\"
                            
                            st.session_state.suggested_path = drive_path
                            st.session_state.folder_path = drive_path
                            st.success(f"✅ Suggested path: {drive_path}")
                            st.rerun()
            else:
                st.warning("⚠️ Hiçbir sürücü bulunamadı!")
                if platform.system() == "Linux":
                    st.info("💡 Docker: Sürücüler mount edilmemiş olabilir")
                else:
                    st.info("💡 Windows: Sürücüler erişilebilir değil")
            
            # Show suggested path and clear button
            if hasattr(st.session_state, 'suggested_path') and st.session_state.suggested_path:
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.info(f"💡 Suggested path: `{st.session_state.suggested_path}`")
                with col2:
                    if st.button("🗑️ Clear", use_container_width=True):
                        del st.session_state.suggested_path
                        if hasattr(st.session_state, 'folder_path'):
                            del st.session_state.folder_path
                        st.rerun()
            
            # Debug information (for developer mode)
            with st.expander("🔧 Debug Info"):
                st.write(f"**Current OS:** {platform.system()}")
                st.write(f"**Environment:** {'Docker Container' if platform.system() == 'Linux' else 'Native System'}")
                st.write(f"**Entered path:** `{folder_path}`")
                
                if folder_path:
                    normalized_path = normalize_windows_path(folder_path)
                    st.write(f"**Normalized path:** `{normalized_path}`")
                    
                    if normalized_path is None:
                        st.write("**Path exists:** ❌ Drive not accessible")
                        st.write("**🔍 Path analysis:** Normalized path is None - drive not mounted")
                    else:
                        st.write(f"**Path exists:** {os.path.exists(normalized_path)}")
                        
                        if os.path.exists(normalized_path):
                            st.write(f"**Is folder:** {os.path.isdir(normalized_path)}")
                            st.write(f"**Is file:** {os.path.isfile(normalized_path)}")
                        else:
                            # Detailed error analysis
                            st.write("**🔍 Detailed path analysis:**")
                            st.write(f"- Path empty: {not normalized_path}")
                            st.write(f"- Path length: {len(normalized_path) if normalized_path else 0}")
                            if normalized_path and len(normalized_path) >= 2:
                                st.write(f"- First char: '{normalized_path[0]}'")
                                st.write(f"- Second char: '{normalized_path[1]}'")
                                if len(normalized_path) >= 3:
                                    st.write(f"- Third char: '{normalized_path[2]}'")
                        
                        # Docker environment check
                        if platform.system() == "Linux":
                            st.write("**🐳 Docker Environment Check:**")
                            # Check if Docker host mounts are available
                            docker_mounts = ['/host/c', '/host/d', '/host/e', '/host/f']
                            available_mounts = []
                            for mount in docker_mounts:
                                if os.path.exists(mount):
                                    available_mounts.append(mount)
                            st.write(f"- Available Docker mounts: {', '.join(available_mounts) if available_mounts else 'None'}")
                            
                            # Check specific drive access
                            if normalized_path and normalized_path.startswith('/host/'):
                                drive_part = normalized_path.split('/')[2] if len(normalized_path.split('/')) > 2 else ''
                                if drive_part:
                                    st.write(f"- Checking Docker mount: /host/{drive_part}")
                                    st.write(f"- Mount exists: {os.path.exists(f'/host/{drive_part}')}")
                        
                        # Windows drive check (native)
                        elif platform.system() == "Windows" and normalized_path and len(normalized_path) >= 2:
                            if normalized_path[1] == ':':
                                drive_letter = normalized_path[0].upper()
                                st.write(f"**💾 Drive check:**")
                                st.write(f"- Drive letter: {drive_letter}")
                                st.write(f"- Drive exists (forward slash): {os.path.exists(drive_letter + ':/')}")
                                backslash_path = drive_letter + ':\\'
                                st.write(f"- Drive exists (backslash): {os.path.exists(backslash_path)}")
                                
                                # List available drives
                                import string
                                available_drives = []
                                for letter in string.ascii_uppercase:
                                    drive_path_forward = f"{letter}:/"
                                    drive_path_backslash = f"{letter}:\\"
                                    if os.path.exists(drive_path_forward) or os.path.exists(drive_path_backslash):
                                        available_drives.append(letter)
                                st.write(f"- Available drives: {', '.join(available_drives)}")
                
                st.write(f"**Session state folder_path:** `{st.session_state.get('folder_path', 'None')}`")
                st.write(f"**Session state suggested_path:** `{st.session_state.get('suggested_path', 'None')}`")
        else:
            # Temporary solution for folder picker
            st.info("💡 Please use manual path input for folder picker feature")
            folder_path = st.text_input(
                "Enter folder path:", 
                placeholder="C:\\Users\\YourName\\Documents",
                help="Enter the full path of the folder to analyze"
            )
        
        st.markdown('<h3>🔍 Filters</h3>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            file_type_filter = st.multiselect(
                "File Types:",
                list(analyzer.file_categories.keys()),
                default=[],
                help="Select file categories to include in analysis"
            )
        
        with col2:
            search_filter = st.text_input(
                "Search term:", 
                placeholder="Enter file or folder name...",
                help="Search for files/folders by name"
            )
            search_filter = search_filter if search_filter.strip() else None
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("🚀 Analyze Folder", type="primary", help="Start folder analysis with current filters", use_container_width=True):
                # Session state'den folder_path'i al veya mevcut değeri kullan
                current_folder_path = folder_path
                if hasattr(st.session_state, 'folder_path') and st.session_state.folder_path:
                    current_folder_path = st.session_state.folder_path
                
                # Yolu normalize et ve doğrula
                current_folder_path = normalize_windows_path(current_folder_path)
                
                # Advanced path validation
                if not current_folder_path:
                    st.error("❌ Please enter a folder path!")
                elif current_folder_path is None:
                    # normalize_windows_path None döndürdüyse sürücü mevcut değil
                    st.error("❌ Drive not found or not accessible!")
                    
                    # Mevcut sürücüleri göster
                    available_drives = []
                    if platform.system() == "Linux":  # Docker environment
                        for drive in 'cdefghijklmnopqrstuvwxyz':
                            mount_path = f'/host/{drive}'
                            if os.path.exists(mount_path):
                                available_drives.append(drive.upper())
                    else:  # Native Windows
                        import string
                        for letter in string.ascii_uppercase:
                            drive_path = f"{letter}:\\"
                            if os.path.exists(drive_path):
                                available_drives.append(letter)
                    
                    if available_drives:
                        st.info(f"💡 Mevcut sürücüler: {', '.join(available_drives)}")
                    else:
                        st.info("💡 Hiçbir sürücü bulunamadı!")
                    
                    if platform.system() == "Linux":
                        st.info("💡 Docker: Sürücüler mount edilmemiş olabilir")
                    else:
                        st.info("💡 Windows: Sürücüler erişilebilir değil")
                elif not os.path.exists(current_folder_path):
                    st.error(f"❌ Folder not found: {current_folder_path}")
                    
                    # Special message for drive access
                    if current_folder_path and len(current_folder_path) >= 2 and current_folder_path[1] == ':':
                        drive_letter = current_folder_path[0].upper()
                        st.error(f"❌ {drive_letter}: drive not found or not accessible!")
                        st.info("💡 This drive may not exist or not be connected")
                        st.info("💡 It could be a USB drive, CD/DVD drive or network drive")
                    
                    st.info("💡 Tip: Make sure to enter the full path (e.g., C:\\Users\\YourName\\Documents)")
                    st.info("💡 On Windows, make sure to write the drive letter and folder name correctly")
                elif not os.path.isdir(current_folder_path):
                    st.error(f"❌ This is not a folder: {current_folder_path}")
                    st.info("💡 Please enter a folder path, not a file path")
                else:
                    try:
                        with st.spinner("📊 Analyzing folder contents..."):
                            active_extensions = []
                            for category in file_type_filter:
                                active_extensions.extend(analyzer.file_categories[category])
                            
                            files_data = analyzer.analyze_folder_contents(
                                current_folder_path,
                                active_extensions if active_extensions else None,
                                None,
                                None,
                                search_filter
                            )
                            
                            st.session_state.files_data = files_data
                            st.session_state.folder_path = current_folder_path
                            st.session_state.analysis_complete = True
                        
                        st.success("✅ Analysis completed!")
                    except PermissionError:
                        st.error("❌ You don't have permission to access this folder!")
                        st.info("💡 Try running as administrator or select a different folder")
                    except Exception as e:
                        st.error(f"❌ Error occurred during analysis: {str(e)}")
                        st.info("💡 Please try a different folder or restart the application")
    
    if hasattr(st.session_state, 'analysis_complete') and st.session_state.analysis_complete:
        # Session state'den güvenli şekilde veri al
        files_data = st.session_state.get('files_data', [])
        folder_path = st.session_state.get('folder_path', 'Unknown')
        
        # Gerekli veriler var mı kontrol et
        if not files_data:
            st.warning("⚠️ Analiz verisi bulunamadı!")
            st.info("💡 Lütfen tekrar analiz yapın")
            return
        
        st.subheader("📊 Analysis Summary")
        
        total_items = len(files_data)
        files_count = len([item for item in files_data if item['Type'] == 'File'])
        folders_count = len([item for item in files_data if item['Type'] == 'Folder'])
        error_count = len([item for item in files_data if item['Type'] == 'Error'])
        total_size_gb = sum(item['Size (GB)'] for item in files_data if item['Type'] in ['File', 'Folder'])
        
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.metric("📁 Total", total_items, help="Total number of items")
        with col2:
            st.metric("📄 Files", files_count, help="Number of files")
        with col3:
            st.metric("📂 Folders", folders_count, help="Number of folders")
        with col4:
            st.metric("❌ Errors", error_count, help="Number of errors")
        with col5:
            st.metric("💾 Size", f"{total_size_gb:.2f} GB", help="Total size in GB")
        
        st.markdown(f"""
        <div class="folder-info max-w-4xl mx-auto w-full">
            <strong>📁 Folder:</strong> {get_folder_name(folder_path)}<br>
            <strong>📅 Date:</strong> {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
        </div>
        """, unsafe_allow_html=True)
        
        tab1, tab2, tab3 = st.tabs([
            "📋 Data", 
            "📈 Charts", 
            "📄 Export"
        ])
        
        with tab1:
            st.subheader("📋 File Analysis Data")
            
            df = pd.DataFrame(files_data)
            
            # DataFrame boş mu kontrol et
            if df.empty:
                st.warning("⚠️ Hiçbir dosya veya klasör bulunamadı!")
                st.info("💡 Bu klasör boş olabilir veya erişim izniniz olmayabilir")
                return
            
            # Gerekli sütunlar var mı kontrol et
            required_columns = ['Name', 'Type', 'Size (GB)', 'Extension', 'Category']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"❌ Eksik sütunlar: {', '.join(missing_columns)}")
                st.info("💡 Analiz verilerinde sorun var")
                return
            
            col1, col2 = st.columns(2)
            with col1:
                type_filter = st.selectbox("Filter by Type:", ["All", "File", "Folder", "Error"], help="Filter by item type")
            with col2:
                # Category sütunu var mı kontrol et
                if 'Category' in df.columns:
                    categories = list(set(df['Category'].unique()))
                    category_filter = st.selectbox("Filter by Category:", ["All"] + categories, help="Filter by file category")
                else:
                    category_filter = "All"
                    st.info("ℹ️ Category filter not available")
            
            filtered_df = df.copy()
            if type_filter != "All":
                filtered_df = filtered_df[filtered_df['Type'] == type_filter]
            if category_filter != "All" and 'Category' in filtered_df.columns:
                filtered_df = filtered_df[filtered_df['Category'] == category_filter]
            
            # Filtrelenmiş DataFrame boş mu kontrol et
            if filtered_df.empty:
                st.warning("⚠️ Seçilen filtrelere uygun dosya/klasör bulunamadı!")
                st.info("💡 Farklı filtreler deneyin")
                return
            
            st.dataframe(
                filtered_df[['Name', 'Type', 'Size (GB)', 'Extension']],
                use_container_width=True,
                hide_index=True,
                height=400
            )
        
        with tab2:
            st.subheader("📈 Analysis Charts")
            
            # DataFrame boş mu kontrol et
            if df.empty:
                st.warning("⚠️ Grafik gösterilemiyor - veri yok!")
                return
            
            # Gerekli sütunlar var mı kontrol et
            required_columns = ['Name', 'Type', 'Size (GB)', 'Extension']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"❌ Grafik gösterilemiyor - eksik sütunlar: {', '.join(missing_columns)}")
                return
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Sadece dosyalar için grafik
                files_df = df[df['Type'] == 'File']
                if not files_df.empty and 'Extension' in files_df.columns:
                    file_types = files_df['Extension'].value_counts()
                    if len(file_types) > 0:
                        fig1 = px.pie(
                            values=file_types.values,
                            names=file_types.index,
                            title="📄 File Types"
                        )
                        fig1.update_layout(height=300)
                        st.plotly_chart(fig1, use_container_width=True)
                    else:
                        st.info("ℹ️ Dosya türü grafiği gösterilemiyor")
                else:
                    st.info("ℹ️ Dosya bulunamadı")
            
            with col2:
                # En büyük öğeler için grafik
                if 'Size (GB)' in df.columns:
                    top_items = df.nlargest(10, 'Size (GB)')
                    if not top_items.empty:
                        fig2 = px.bar(
                            top_items,
                            x='Name',
                            y='Size (GB)',
                            color='Type',
                            title="📊 Top 10 Largest Items"
                        )
                        fig2.update_xaxes(tickangle=45)
                        fig2.update_layout(height=300)
                        st.plotly_chart(fig2, use_container_width=True)
                    else:
                        st.info("ℹ️ Boyut grafiği gösterilemiyor")
                else:
                    st.info("ℹ️ Boyut bilgisi bulunamadı")
        
        with tab3:
            st.subheader("📄 Export Options")
            
            # Export için veri kontrolü
            if df.empty:
                st.warning("⚠️ Export yapılamıyor - veri yok!")
                return
            
            required_columns = ['Name', 'Type', 'Size (GB)', 'Extension', 'Category']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"❌ Export yapılamıyor - eksik sütunlar: {', '.join(missing_columns)}")
                return
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("📊 Excel", help="Export data to Excel format"):
                    if not df.empty:
                        df_export = pd.DataFrame(files_data)
                        
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                            
                            df_export.to_excel(writer, sheet_name='Detailed Analysis', index=False)
                            
                            workbook = writer.book
                            worksheet = writer.sheets['Detailed Analysis']
                            
                            title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
                            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
                            normal_font = Font(name='Segoe UI', size=10)
                            
                            title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                            header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                            folder_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                            file_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                            error_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                            
                            center_alignment = Alignment(horizontal='center', vertical='center')
                            left_alignment = Alignment(horizontal='left', vertical='center')
                            
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            
                            worksheet.insert_rows(1, 3)
                            worksheet.merge_cells('A1:D1')
                            title_cell = worksheet['A1']
                            folder_display_name = get_folder_name(folder_path)
                            title_cell.value = f"📁 File Size Analysis - {folder_display_name}"
                            title_cell.font = title_font
                            title_cell.fill = title_fill
                            title_cell.alignment = center_alignment
                            title_cell.border = thin_border
                            
                            worksheet.merge_cells('A2:D2')
                            subtitle_cell = worksheet['A2']
                            subtitle_cell.value = f"📊 Analysis Date: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
                            subtitle_cell.font = Font(name='Segoe UI', size=9, color='666666')
                            subtitle_cell.alignment = center_alignment
                            
                            headers = ['Name', 'Type', 'Size (GB)', 'Extension']
                            for col, header in enumerate(headers, 1):
                                cell = worksheet.cell(row=4, column=col)
                                cell.value = header
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = center_alignment
                                cell.border = thin_border
                            
                            # Tüm sayfa için biçimlendirme
                            for row_num in range(1, len(df_export) + 5):
                                for col_num in range(1, 5):
                                    cell = worksheet.cell(row=row_num, column=col_num)
                                    
                                    # Başlık satırları (1-3)
                                    if row_num <= 3:
                                        if row_num == 1:
                                            cell.font = title_font
                                            cell.fill = title_fill
                                            cell.alignment = center_alignment
                                        elif row_num == 2:
                                            cell.font = Font(name='Segoe UI', size=9, color='666666')
                                            cell.alignment = center_alignment
                                        elif row_num == 3:
                                            cell.font = Font(name='Segoe UI', size=10, color='666666')
                                            cell.alignment = center_alignment
                                    # Başlık satırı (4)
                                    elif row_num == 4:
                                        cell.font = header_font
                                        cell.fill = header_fill
                                        cell.alignment = center_alignment
                                    # Veri satırları (5+)
                                    else:
                                        cell.font = normal_font
                                        cell.alignment = left_alignment
                                        
                                        row_type = worksheet.cell(row=row_num, column=2).value
                                        if row_type == 'Folder':
                                            cell.fill = folder_fill
                                        elif row_type == 'File':
                                            cell.fill = file_fill
                                        elif row_type == 'Error':
                                            cell.fill = error_fill
                                    
                                    # Tüm hücrelere border ekle
                                    cell.border = thin_border
                            
                            for column in worksheet.columns:
                                max_length = 0
                                column_letter = None
                                for cell in column:
                                    if cell.coordinate in worksheet.merged_cells:
                                        continue
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                            column_letter = cell.column_letter
                                    except:
                                        pass
                                if column_letter:
                                    adjusted_width = min(max_length + 2, 50)
                                    worksheet.column_dimensions[column_letter].width = adjusted_width
                            
                            # Özet sayfası
                            summary_data = {
                                'Metric': [
                                    'Total Items',
                                    'File Count',
                                    'Folder Count',
                                    'Error Count',
                                    'Total Size (GB)',
                                    'Largest File',
                                    'Smallest File',
                                    'Average File Size (GB)',
                                    'Analyzed Folders'
                                ],
                                'Value': [
                                    total_items,
                                    files_count,
                                    folders_count,
                                    error_count,
                                    round(total_size_gb, 2),
                                    max([item for item in files_data if item['Type'] == 'File'], 
                                        key=lambda x: x['Size (GB)'])['Name'] if files_count > 0 else 'None',
                                    min([item for item in files_data if item['Type'] == 'File'], 
                                        key=lambda x: x['Size (GB)'])['Name'] if files_count > 0 else 'None',
                                    round(total_size_gb / files_count, 2) if files_count > 0 else 0,
                                    len(df_export['Name'].unique())
                                ]
                            }
                            
                            summary_df = pd.DataFrame(summary_data)
                            summary_df.to_excel(writer, sheet_name='General Summary', index=False)
                            
                            # Özet sayfasını formatla
                            summary_worksheet = writer.sheets['General Summary']
                            summary_worksheet.merge_cells('A1:B1')
                            title_cell2 = summary_worksheet['A1']
                            title_cell2.value = f"📊 File Analysis Summary - {folder_display_name}"
                            title_cell2.font = title_font
                            title_cell2.fill = title_fill
                            title_cell2.alignment = center_alignment
                            title_cell2.border = thin_border
                            
                            # Özet başlıkları
                            headers2 = ['Metric', 'Value']
                            for col, header in enumerate(headers2, 1):
                                cell = summary_worksheet.cell(row=3, column=col)
                                cell.value = header
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = center_alignment
                                cell.border = thin_border
                            
                            # Özet verileri
                            for row_idx in range(4, len(summary_df) + 4):
                                for col_idx in range(1, 3):
                                    cell = summary_worksheet.cell(row=row_idx, column=col_idx)
                                    cell.font = normal_font
                                    cell.alignment = left_alignment
                                    cell.border = thin_border
                                    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                            
                            file_types = df_export[df_export['Type'] == 'File']['Extension'].value_counts().reset_index()
                            file_types.columns = ['File Extension', 'Count']
                            file_types.to_excel(writer, sheet_name='File Types', index=False)
                            
                            types_worksheet = writer.sheets['File Types']
                            types_worksheet.merge_cells('A1:B1')
                            title_cell3 = types_worksheet['A1']
                            title_cell3.value = f"📄 File Types Analysis - {folder_display_name}"
                            title_cell3.font = title_font
                            title_cell3.fill = title_fill
                            title_cell3.alignment = center_alignment
                            title_cell3.border = thin_border
                            
                            headers3 = ['File Extension', 'Count']
                            for col, header in enumerate(headers3, 1):
                                cell = types_worksheet.cell(row=3, column=col)
                                cell.value = header
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = center_alignment
                                cell.border = thin_border
                            
                            for row_idx in range(4, len(file_types) + 4):
                                for col_idx in range(1, 3):
                                    cell = types_worksheet.cell(row=row_idx, column=col_idx)
                                    cell.font = normal_font
                                    cell.alignment = left_alignment
                                    cell.border = thin_border
                                    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                            
                            folder_summary = df_export.groupby('Name').agg({
                                'Type': 'count',
                                'Size (GB)': 'sum'
                            }).reset_index()
                            folder_summary.columns = ['Name', 'Item Count', 'Total Size (GB)']
                            folder_summary['Total Size (GB)'] = folder_summary['Total Size (GB)'].round(2)
                            folder_summary = folder_summary.sort_values('Total Size (GB)', ascending=False)
                            folder_summary.to_excel(writer, sheet_name='Folder Summary', index=False)
                            
                            folder_worksheet = writer.sheets['Folder Summary']
                            folder_worksheet.merge_cells('A1:C1')
                            title_cell4 = folder_worksheet['A1']
                            title_cell4.value = f"📂 Folder Summary - {folder_display_name}"
                            title_cell4.font = title_font
                            title_cell4.fill = title_fill
                            title_cell4.alignment = center_alignment
                            title_cell4.border = thin_border
                            
                            headers4 = ['Name', 'Item Count', 'Total Size (GB)']
                            for col, header in enumerate(headers4, 1):
                                cell = folder_worksheet.cell(row=3, column=col)
                                cell.value = header
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = center_alignment
                                cell.border = thin_border
                            
                            for row_idx in range(4, len(folder_summary) + 4):
                                for col_idx in range(1, 4):
                                    cell = folder_worksheet.cell(row=row_idx, column=col_idx)
                                    cell.font = normal_font
                                    cell.alignment = left_alignment
                                    cell.border = thin_border
                                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                            
                            if 'Sheet' in workbook.sheetnames:
                                workbook.remove(workbook['Sheet'])
                        
                        output.seek(0)
                        
                        st.download_button(
                            label="📥 Download Excel File",
                            data=output.getvalue(),
                            file_name=f"detailed_file_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            
            with col2:
                if st.button("📄 CSV", help="Export data to CSV format"):
                    if len(files_data) > 0:
                        df_export = pd.DataFrame(files_data)
                        csv = df_export.to_csv(index=False)
                        
                        st.download_button(
                            label="📥 Download CSV File",
                            data=csv,
                            file_name=f"file_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv"
                        )
            
            with col3:
                if st.button("📊 Package", help="Export data package with charts"):
                    if len(files_data) > 0:
                        zip_buffer = BytesIO()
                        
                        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                            json_data = json.dumps(files_data, indent=2, default=str)
                            zip_file.writestr('analysis_data.json', json_data)
                            
                            summary_text = f"""
File Analysis Summary
====================
Folder: {folder_path}
Date: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}

Total Items: {total_items}
Files: {files_count}
Folders: {folders_count}
Errors: {error_count}
Total Size: {total_size_gb:.2f} GB
                            """
                            zip_file.writestr('summary.txt', summary_text)
                        
                        zip_buffer.seek(0)
                        
                        st.download_button(
                            label="📥 Download Data Package",
                            data=zip_buffer.getvalue(),
                            file_name=f"file_analysis_package_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                            mime="application/zip"
                        )
 
analyzer = FileSizeAnalyzerWeb()

if __name__ == "__main__":
    main() 